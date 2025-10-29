'''
Script to parse constituency wise data reports (usually report nos. 32, 33).
Pass .xlsx files as input - ECI provides .xls files, must be converted to .xlsx files using Excel to use with this script.
'''

from openpyxl import load_workbook
import json
from enum import Enum
import string

# enum to track section of the summary report being parsed currently
class CurrentSection(Enum):
    STATE_UT = "State/UT"
    CANDIDATES = "Candidates"
    ELECTORS = "Electors"
    VOTERS = "Voters"
    VOTES = "Votes"
    POLLING_STATION = "Polling_Station"
    DATES = "Dates"
    RESULT = "Result"
    NONE = "None"

'''
Function to format constituency names
'''
def format_constituency_name(name):
    parts = name.split('-')
    parts = [string.capwords(part.strip()) for part in parts]
    return '-'.join(parts).replace('&', 'and')

'''
Function to parse a sheet in the summary report, each sheet corresponds to one constituency.
In the first half of the for loop, check and define the section being processed currently. In the second half, parse the rows according to the current section as assigned in the first half (except for single-row-sections which are parsed in the first half).
'''
def parse_summary_sheet(sheet):
    data = {
        "ID": sheet.title,
        "Constituency": None,
        "State_UT": None,
        "Category": None,
        "Candidates": {},
        "Electors": {},
        "Voters": {},
        "Votes": {},
        "Polling_Station": {},
        "Dates": [],
        "Result": {}
    }
    
    current_section = CurrentSection.NONE
    
    for row in sheet.iter_rows(values_only=True):
        # skip empty rows
        if not any(row):
            continue 
    
        # detect/assign sections
        if "State/UT" in str(row[0]):
            current_section = CurrentSection.STATE_UT
            # state/UT and constituency data are specified in the same row, parse them
            data["State_UT"] = row[1].split('-')[0]
            data["Constituency"] = format_constituency_name('-'.join(row[3].split('-')[:-1]))
            data["Category"] = row[3].split('-')[-1] # last field specifies type of constituency
        elif "CANDIDATES" in str(row[0]):
            current_section = CurrentSection.CANDIDATES
            continue
        elif "ELECTORS" in str(row[0]):
            current_section = CurrentSection.ELECTORS
            continue
        elif "VOTERS" in str(row[0]):
            current_section = CurrentSection.VOTERS
            continue
        elif "VOTES" in str(row[0]):
            current_section = CurrentSection.VOTES
            continue
        elif "POLLING STATION" in str(row[0]):
            current_section = CurrentSection.POLLING_STATION
            continue
        elif "DATES" in str(row[0]):
            current_section = CurrentSection.DATES
            # dates of elections are specified in the same row, parse the dates
            idx = 3
            while row[idx]:
                data[current_section.value].append(str(row[idx]))
                idx += 1
        elif "RESULT" in str(row[0]):
            current_section = CurrentSection.RESULT
            continue
        
        # parse sections
        # 'CANDIDATES', 'ELECTORS' and 'VOTERS' sections have similar structure, use same logic
        if current_section == CurrentSection.CANDIDATES or current_section == CurrentSection.ELECTORS or current_section == CurrentSection.VOTERS:
            data[current_section.value][str(row[1])] = {"Men": row[3], "Women": row[4], "Third_Gender": row[5], "Total": row[6]}
        elif current_section == CurrentSection.VOTES:
            data[current_section.value][str(row[1])] = row[6]
        elif current_section == CurrentSection.POLLING_STATION:
            if str(row[1]) == "Dates(s) of Re-Poll if Any":
                dates = []
                idx = 4
                while row[idx]:
                    dates.append(str(row[idx]))
                    idx += 1
                if dates: # verify that there was/were repoll(s), avoids empty entries
                    data[current_section.value][str(row[1])] = dates
            elif str(row[1]) == "Numbers Of Polling Stations where Re-Poll was Order":
                if "Dates(s) of Re-Poll if Any" in data[current_section.value].keys(): # same check as dates, avoids empty/unnecessary entries
                    data[current_section.value][str(row[1])] = row[6]
            else:
                data[current_section.value][str(row[1])] = row[3]
                data[current_section.value][str(row[4])] = row[6]    
        elif current_section == CurrentSection.RESULT:
            if str(row[1]) != "Margin":
                data[current_section.value][str(row[1])] = {"Party": row[3], "Candidates": row[4], "Votes": row[6]}
            else:
                data[current_section.value][str(row[1])] = row[3]
                current_section = CurrentSection.NONE

    return data

'''
Function to parse the detailed result report which contains candidate-wise data for each constituency.
Each row corresponds to one candidate.
'''
def parse_detailed_sheet(sheet, ids):
    candidates = dict()
    rows = list(sheet.iter_rows(values_only=True))
    l1_fields = list(rows[1])
    l2_fields = list(rows[2])

    # clean the headers
    # TODO: optimize this
    for i in range(len(l1_fields)):
        if l1_fields[i]:
            l1_fields[i] = l1_fields[i].replace('\n', ' ')
    for i in range(len(l2_fields)):
        if l2_fields[i]:
            l2_fields[i] = l2_fields[i].replace('\n', ' ')

    # manually copy over some l1 fields wherever needed
    # TODO: find a way to automate this
    l1_fields[11] = l1_fields[10]
    l1_fields[12] = l1_fields[10]
    l1_fields[14] = l1_fields[13]

    for row in rows[3:]:
        # skip empty and irrelevant rows
        if not row[2]: # check arbitrary field, this is non-empty for rows containing data
            continue

        state = row[0]
        constituency = format_constituency_name(row[1])
        constituency_id = 'NA'
        for key in ids.keys():
            if ids[key]['State_UT'] == state and ids[key]['Constituency'] == constituency:
                constituency_id = key
                break
        if constituency_id == 'NA':
            print('Not found', constituency, state)

        candidate = {}
        for field in range(2, len(row)):
            if row[field]:
                if l1_fields[field]:
                    try:
                        candidate[l1_fields[field]][l2_fields[field]] = row[field]
                    except:
                        candidate[l1_fields[field]] = {l2_fields[field]: row[field]}
                else:
                    candidate[l2_fields[field]] = row[field]
        try:
            candidates[constituency_id].append(candidate)
        except:
            candidates[constituency_id] = [candidate]
    return candidates


summary_file_path = '/home/gmangipu/uni/courses/cs699/project/scripts/data/2024/to_parse/32-Constituency-Data-Summery-Report.xlsx'
detailed_file_path = '/home/gmangipu/uni/courses/cs699/project/scripts/data/2024/to_parse/33-Constituency-Wise-Detailed-Result.xlsx'
out_path = 'parsed/2024.json'

# get constituency-wise summary data
wb = load_workbook(summary_file_path)
parsed = []
for s in wb.sheetnames:
    sheet = wb[s]
    parsed.append(parse_summary_sheet(sheet))

# extract ID of each constituency, useful to parse other reports
ids = {}
for c in parsed:
    ids[c['ID']] = {'State_UT': c['State_UT'], 'Constituency': c['Constituency']}

# get constituency-wise candidate data
wb = load_workbook(detailed_file_path)
candidates = parse_detailed_sheet(wb.active, ids)

# merge candidate data into parsed summary data
for constituency in parsed:
    constituency['Candidates'] = candidates[constituency['ID']]

# dump the parsed data in json format
with open(out_path, 'w') as f:
    json_string = json.dumps(parsed, indent=4)
    json_string = json_string.replace('"=(0)"', '0') # replace formula string with absolute value
    f.write(json_string)