"""
Combined script to parse constituency-wise data reports for 2009 (PDF),
2014 (XLSX), 2019 (XLSX), and 2024 (XLSX).

"""

import json
from enum import Enum
import string
import sys
import re
import os
from collections import defaultdict

try:
    import pdfplumber
except ImportError:
    # This check is kept for the 2009 PDF parser
    pass 

try:
    from openpyxl import load_workbook
except ImportError:
    # This check is kept for the 2014/2019/2024 XLSX parsers
    pass


# --------------------------------------------------------------------------
# CONFIGURATION
# --------------------------------------------------------------------------

JOBS_CONFIG = {
    2009: {
        "summary_path": "converted_xlsx_reports/Summary_details/Constituency Data Summary.pdf",
        "detailed_path": "converted_xlsx_reports/Summary_details/Constituency Wise Detailed Result.pdf",
        "output_path": "parsed_new/parsed_new_2009.json",
        "parser_type": "PDF"
    },
    2014: {
        "summary_path": "converted_xlsx_reports/Summary_details/2014_summary.xlsx",
        "detailed_path": "converted_xlsx_reports/Summary_details/2014_detailed.xlsx",
        "output_path": "parsed_new/parsed_new_2014.json",
        "parser_type": "XLSX"
    },
    2019: {
        "summary_path": 'converted_xlsx_reports/election_data_2019_xlsx/32. Constituency Data Summary Report.xlsx',
        "detailed_path": 'converted_xlsx_reports/election_data_2019_xlsx/33. Constituency Wise Detailed Result.xlsx',
        "output_path": 'parsed_new/parsed_new_2019.json',
        "parser_type": "XLSX"
    },
    2024: {
        "summary_path": 'converted_xlsx_reports/election_data_2024_xlsx/32-Constituency-Data-Summery-Report.xlsx',
        "detailed_path": 'converted_xlsx_reports/election_data_2024_xlsx/33-Constituency-Wise-Detailed-Result.xlsx',
        "output_path": 'parsed_new/parsed_new_2024.json',
        "parser_type": "XLSX"
    },
}

# --------------------------------------------------------------------------
# COMMON ENUMS AND MAPS
# --------------------------------------------------------------------------

class CurrentSection(Enum):
    STATE_UT = "State/UT"
    SUMMARY_CANDIDATE_STATS = "Summary_Candidate_Stats"
    ELECTORS = "Electors"
    VOTERS = "Voters"
    VOTES = "Votes"
    POLLING_STATION = "Polling_Station"
    DATES = "Dates"
    RESULT = "Result"
    NONE = "None"

STATE_UT_MAP_2009 = {
    "S01": "Andhra Pradesh", "S02": "Arunachal Pradesh", "S03": "Assam", "S04": "Bihar", 
    "S05": "Goa", "S06": "Gujarat", "S07": "Haryana", "S08": "Himachal Pradesh", 
    "S09": "Jammu & Kashmir", "S10": "Karnataka", "S11": "Kerala", "S12": "Madhya Pradesh", 
    "S13": "Maharashtra", "S14": "Manipur", "S15": "Meghalaya", "S16": "Mizoram", 
    "S17": "Nagaland", "S18": "Orissa", "S19": "Punjab", "S20": "Rajasthan", 
    "S21": "Sikkim", "S22": "Tamil Nadu", "S23": "Tripura", "S24": "Uttar Pradesh", 
    "S25": "West Bengal", "S26": "Chhattisgarh", "S27": "Jharkhand", "S28": "Uttarakhand", 
    "U01": "Andaman & Nicobar Islands", "U02": "Chandigarh", "U03": "Dadra & Nagar Haveli", 
    "U04": "Daman & Diu", "U05": "National Capital Territory of Delhi", "U06": "Lakshadweep", 
    "U07": "Puducherry",
}

STATE_NAME_CORRECTIONS = {
    "andhra prade": "Andhra Pradesh",
    "orissa": "Odisha",
    "chhattisgarh": "Chhattisgarh",
    "nct of delhi": "NCT OF Delhi",
    "telangana": "Telangana"
}

# --------------------------------------------------------------------------
# COMMON HELPERS (for 2009, 2014, 2019, 2024)
# --------------------------------------------------------------------------

def clean_value(value):
    if isinstance(value, str):
        value = value.replace(u'\xa0', ' ').strip()
        if value.startswith("=(") and value.endswith(")"):
            try:
                return int(value[2:-1])
            except ValueError:
                pass 
    return value

def format_constituency_name(name):
    if not name or not isinstance(name, str):
        return "Unknown"
    
    name_cleaned = str(name).replace(u'\xa0', ' ').strip()
    name_cleaned = re.sub(r"^\s*[\d\s-]+\s*", "", name_cleaned)
    name_cleaned = re.sub(r"\s*\((SC|ST)\)\s*", " ", name_cleaned, flags=re.I).strip()
    name_cleaned = re.sub(r"-(SC|ST)-?\d*$", "", name_cleaned, flags=re.I)
    name_cleaned = re.sub(r"\s*-\s*\d+\s*$", "", name_cleaned)
    name_cleaned = re.sub(r"-Gen$", "", name_cleaned, flags=re.I)

    parts = name_cleaned.split('-')
    parts = [string.capwords(part.strip()) for part in parts if part.strip()]
    return '-'.join(parts).replace('&', 'and')

def safe_int(value):
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '').replace('-', '0').replace('N/A', '0')
        if value.startswith('(') and value.endswith(')'):
            value = value[1:-1]
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0

def safe_float(value):
    if isinstance(value, (float, int)):
        return float(value)
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '').replace('-', '0.0').replace('N/A', '0.0')
        if value.startswith('(') and value.endswith(')'):
            value = value[1:-1]
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

# --------------------------------------------------------------------------
# 2024-Compliant Template Helpers
# --------------------------------------------------------------------------
def get_empty_gender_obj(default_val=0):
    val = 0 if default_val == 0 else None
    return {"Men": val, "Women": val, "Third_Gender": 0, "Total": default_val}

def get_2024_voters_template():
    return {
        "General": get_empty_gender_obj(0),
        "OverSeas": get_empty_gender_obj(0),
        "Proxy": get_empty_gender_obj(0),
        "Postal": get_empty_gender_obj(0),
        "Total": get_empty_gender_obj(0),
        "Votes Not Counted From CU(s) as Per ECI Instructions": get_empty_gender_obj(0),
        "POLLING PERCENTAGE": {"Men": None, "Women": None, "Third_Gender": None, "Total": 0.0}
    }

def get_2024_electors_template():
    return {
        "General": get_empty_gender_obj(0),
        "OverSeas": get_empty_gender_obj(0), 
        "Service": get_empty_gender_obj(0),
        "Total": get_empty_gender_obj(0)
    }

def get_2024_votes_template():
    return {
        "Total Votes Polled On EVM": 0,
        "Total Deducted Votes From EVM": 0,
        "Total Valid Votes polled on EVM": 0,
        "Postal Votes Counted": 0,
        "Postal Votes Deducted": 0,
        "Valid Postal Votes": 0,
        "Total Valid Votes Polled": 0,
        "Test Votes polled On EVM": 0,
        "Votes Polled for 'NOTA'(Including Postal)": 0, 
        "Tendered Votes": 0,
    }


# --------------------------------------------------------------------------
# 2009 PDF PARSERS
# --------------------------------------------------------------------------
def parse_2009_summary_pdf(pdf_path):
    print("\n--- Parsing 2009 Summary PDF (Report 32) ---")
    if 'pdfplumber' not in sys.modules:
        print("Error: 'pdfplumber' not installed. Skipping 2009 parsing.")
        return []

    all_constituency_data = []

    state_re = re.compile(r"State/UT\s*:\s*([A-Z\d]+)", re.I)
    const_re = re.compile(r"Constituency\s*:\s*([^\n\(]+)", re.I)
    id_re = re.compile(r"No\.\s*:\s*(\d+)", re.I)
    
    cand_nominated_re = re.compile(r"1\.\s*NOMINATED\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    cand_rejected_re = re.compile(r"2\.\s*NOMINATION REJECTED\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    cand_withdrawn_re = re.compile(r"3\.\s*WITHDRAWN\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    cand_contested_re = re.compile(r"4\.\s*CONTESTED\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    cand_forfeited_re = re.compile(r"5\.\s*FORFEITED DEPOSIT\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    elec_general_re = re.compile(r"II\.\s*ELECTORS\s*1\.\s*GENERAL\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I | re.DOTALL)
    elec_service_re = re.compile(r"2\.\s*SERVICE\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    elec_total_re = re.compile(r"3\.\s*TOTAL\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I)
    voters_general_re = re.compile(r"III\.\s*VOTERS\s*1\.\s*GENERAL\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)", re.I | re.DOTALL)
    voters_proxy_re = re.compile(r"2\.\s*PROXY\s+([\d-]+)", re.I)
    voters_postal_re = re.compile(r"3\.\s*POSTAL\s+([\d-]+)", re.I)
    voters_total_re = re.compile(r"4\.\s*TOTAL\s+([\d-]+)", re.I)
    polling_percent_re = re.compile(r"III\(A\)\.\s*POLLING PERCENTAGE\s*([\d\.]+)", re.I)
    votes_rejected_re = re.compile(r"1\.\s*REJECTED VOTES \(POSTAL\)\s*([\d-]+)", re.I)
    votes_not_retrieved_re = re.compile(r"2\.\s*VOTES NOT RETREIVED FROM EVM\s*([\d-]+)", re.I)
    votes_valid_re = re.compile(r"3\.\s*TOTAL VALID VOTES POLLED\s*([\d-]+)", re.I)
    votes_tendered_re = re.compile(r"4\. \s*TENDERED VOTES\s*([\d-]+)", re.I)
    ps_number_re = re.compile(r"V\.\s*POLLING STATIONS\s*NUMBER\s*(\d+)", re.I | re.DOTALL)
    ps_avg_re = re.compile(r"AVERAGE ELECTORS PER POLLING STATION\s*(\d+)", re.I)
    dates_polling_re = re.compile(r"POLLING\s+([\d-]+)", re.I)
    dates_counting_re = re.compile(r"COUNTING\s+([\d-]+)", re.I)
    dates_decl_re = re.compile(r"DECLARATION\s+([\d-]+)", re.I)

    def find_groups(regex, text, num_groups=1):
        match = regex.search(text)
        default = ["0"] * num_groups
        if not match: return default
        return [match.group(i+1).strip() for i in range(num_groups)]

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text(x_tolerance=1, y_tolerance=3)
                if not text: continue

                data = {
                    "ID": None, "Constituency": None, "State_UT": None, "Category": None,
                    "Candidates": [], "Summary_Candidate_Stats": {},
                    "Electors": get_2024_electors_template(), "Voters": get_2024_voters_template(),
                    "Votes": get_2024_votes_template(),
                    "Polling_Station": {"Number": 0, "Average Electors Per Polling": 0},
                    "Dates": [],
                    "Result": {"Winner": {"Party": None, "Candidates": None, "Votes": 0},
                               "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0},
                               "Margin": 0}
                }

                state_match = state_re.search(text)
                const_match = const_re.search(text)
                id_match = id_re.search(text)

                if state_match and const_match and id_match:
                    state_code = state_match.group(1).strip().upper()
                    data["State_UT"] = STATE_UT_MAP_2009.get(state_code, state_code)
                    data["ID"] = f"{state_code}-{id_match.group(1).strip()}"
                    constituency_full = const_match.group(1).strip()
                    cat_match = re.search(r"\((ST|SC)\)", text, re.I)
                    data["Category"] = cat_match.group(1).upper() if cat_match else "GENERAL"
                    data["Constituency"] = format_constituency_name(constituency_full)
                    
                    nom = find_groups(cand_nominated_re, text, 3)
                    rej = find_groups(cand_rejected_re, text, 3)
                    wd = find_groups(cand_withdrawn_re, text, 3)
                    con = find_groups(cand_contested_re, text, 3)
                    data["Summary_Candidate_Stats"]["Nominated"] = {"Men": safe_int(nom[0]), "Women": safe_int(nom[1]), "Third_Gender": 0, "Total": safe_int(nom[2])}
                    data["Summary_Candidate_Stats"]["Nomination Rejected"] = {"Men": safe_int(rej[0]), "Women": safe_int(rej[1]), "Third_Gender": 0, "Total": safe_int(rej[2])}
                    data["Summary_Candidate_Stats"]["Withdrawn"] = {"Men": safe_int(wd[0]), "Women": safe_int(wd[1]), "Third_Gender": 0, "Total": safe_int(wd[2])}
                    data["Summary_Candidate_Stats"]["Contested"] = {"Men": safe_int(con[0]), "Women": safe_int(con[1]), "Third_Gender": 0, "Total": safe_int(con[2])}
                    
                    gen_e = find_groups(elec_general_re, text, 3)
                    ser_e = find_groups(elec_service_re, text, 3)
                    tot_e = find_groups(elec_total_re, text, 3)
                    data["Electors"]["General"] = {"Men": safe_int(gen_e[0]), "Women": safe_int(gen_e[1]), "Third_Gender": 0, "Total": safe_int(gen_e[2])}
                    data["Electors"]["Service"] = {"Men": safe_int(ser_e[0]), "Women": safe_int(ser_e[1]), "Third_Gender": 0, "Total": safe_int(ser_e[2])}
                    data["Electors"]["Total"] = {"Men": safe_int(tot_e[0]), "Women": safe_int(tot_e[1]), "Third_Gender": 0, "Total": safe_int(tot_e[2])}

                    gen_v = find_groups(voters_general_re, text, 3)
                    prox_v = find_groups(voters_proxy_re, text, 1)
                    post_v = find_groups(voters_postal_re, text, 1)
                    tot_v = find_groups(voters_total_re, text, 1)
                    poll_pct = find_groups(polling_percent_re, text, 1)
                    data["Voters"]["General"] = {"Men": safe_int(gen_v[0]), "Women": safe_int(gen_v[1]), "Third_Gender": 0, "Total": safe_int(gen_v[2])}
                    data["Voters"]["Proxy"]["Total"] = safe_int(prox_v[0])
                    data["Voters"]["Postal"]["Total"] = safe_int(post_v[0])
                    data["Voters"]["Total"]["Total"] = safe_int(tot_v[0])
                    data["Voters"]["POLLING PERCENTAGE"]["Total"] = safe_float(poll_pct[0])

                    rej_v_postal_text = find_groups(votes_rejected_re, text, 1)
                    not_ret_v = find_groups(votes_not_retrieved_re, text, 1)
                    valid_v = find_groups(votes_valid_re, text, 1)
                    tend_v = find_groups(votes_tendered_re, text, 1)

                    total_rejected_votes = safe_int(rej_v_postal_text[0])
                    total_votes_polled = safe_int(tot_v[0])
                    total_valid_votes = safe_int(valid_v[0])
                    
                    if total_votes_polled - total_valid_votes != total_rejected_votes:
                        total_rejected_votes = total_votes_polled - total_valid_votes
                    
                    data["Votes"]["Postal Votes Counted"] = safe_int(post_v[0])
                    data["Votes"]["Total Votes Polled On EVM"] = safe_int(tot_v[0]) - safe_int(post_v[0])
                    data["Votes"]["Total Valid Votes Polled"] = total_valid_votes
                    data["Votes"]["Tendered Votes"] = safe_int(tend_v[0])
                    data["Votes"]["Total Deducted Votes From EVM"] = safe_int(not_ret_v[0])
                    postal_deducted = total_rejected_votes - data["Votes"]["Total Deducted Votes From EVM"]
                    
                    if postal_deducted < 0: postal_deducted = 0

                    if postal_deducted > data["Votes"]["Postal Votes Counted"]:
                        data["Votes"]["Total Deducted Votes From EVM"] = total_rejected_votes
                        data["Votes"]["Postal Votes Deducted"] = 0
                    else:
                        data["Votes"]["Postal Votes Deducted"] = postal_deducted
                    
                    data["Votes"]["Valid Postal Votes"] = data["Votes"]["Postal Votes Counted"] - data["Votes"]["Postal Votes Deducted"]
                    data["Votes"]["Total Valid Votes polled on EVM"] = data["Votes"]["Total Votes Polled On EVM"] - data["Votes"]["Total Deducted Votes From EVM"]
                    
                    if data["Votes"]["Total Valid Votes polled on EVM"] < 0:
                        data["Votes"]["Total Valid Votes polled on EVM"] = 0
                    
                    if data["Votes"]["Total Valid Votes polled on EVM"] + data["Votes"]["Valid Postal Votes"] != data["Votes"]["Total Valid Votes Polled"]:
                         data["Votes"]["Total Valid Votes polled on EVM"] = data["Votes"]["Total Valid Votes Polled"] - data["Votes"]["Valid Postal Votes"]

                    ps_num = find_groups(ps_number_re, text, 1)
                    ps_a = find_groups(ps_avg_re, text, 1)
                    data["Polling_Station"]["Number"] = safe_int(ps_num[0])
                    data["Polling_Station"]["Average Electors Per Polling"] = safe_int(ps_a[0])

                    poll_d = find_groups(dates_polling_re, text, 1)
                    decl_d = find_groups(dates_decl_re, text, 1)
                    if poll_d[0] != "0": data["Dates"].append(poll_d[0])
                    if decl_d[0] != "0": data["Dates"].append(decl_d[0])

                    all_constituency_data.append(data)
    except Exception as e:
        print(f"Error opening/parsing 2009 summary PDF: {e}")
        return []

    print(f"--- 2009 Summary PDF parsing complete. Found {len(all_constituency_data)} entries. ---")
    return all_constituency_data

def parse_2009_detailed_pdf(pdf_path, ids_map):
    print("\n--- Parsing 2009 Detailed PDF (Report 33) ---")
    if 'pdfplumber' not in sys.modules: return {}

    state_to_const_map = {}
    for full_id, details in ids_map.items():
        state_upper = details["State_UT"].upper()
        const_upper = details["Constituency"].upper()
        if state_upper not in state_to_const_map: state_to_const_map[state_upper] = {}
        state_to_const_map[state_upper][const_upper] = full_id

    candidates_by_constituency = {}
    current_constituency_id = None
    current_total_electors = 0
    current_state_name = None 
    
    alternate_state_name_map = {
        "DELHI": "NATIONAL CAPITAL TERRITORY OF DELHI", "NCT OF DELHI": "NATIONAL CAPITAL TERRITORY OF DELHI",
        "CHHATTISGARH": "CHHATTISGARH", "CHATTISGARH": "CHHATTISGARH", "CHHATISGARH": "CHHATTISGARH",  
    }
    for name in STATE_UT_MAP_2009.values(): alternate_state_name_map[name.upper()] = name.upper()

    anchor_regex = re.compile(r"([MF])\s+(\d+)\s+([A-Z]{2,3})", re.I)
    normal_const_regex = re.compile(r"CONSTITUENCY\s*:\s*(\d+)?\s*\.?\s*([A-Za-z&\-\s]{3,}[^\(]*?)(?:\((ST|SC)\))?", re.IGNORECASE)
    reverse_const_regex = re.compile(r"([A-Za-z&\-\s]{3,})\s+CONSTITUENCY\s*:", re.IGNORECASE)

    def normalize_name(name: str) -> str:
        if not name: return ""
        name = re.sub(r"\s+", " ", name.strip())
        name = re.sub(r"[^A-Za-z&\-\s]", "", name)
        name = name.replace("’", "'").replace("NAGARH", "NAGAR").replace("UDHAMSINGH", "UDHAMSINGH NAGAR").replace("NAGA", "NAGAR").replace("ISLAND", "ISLANDS")
        return format_constituency_name(name)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text(x_tolerance=1, y_tolerance=3)
                if not text: continue
                lines = [re.sub(r"\s+", " ", ln.strip()) for ln in text.split("\n") if ln.strip()]

                for i, line in enumerate(lines):
                    line_upper_stripped = line.strip().upper()
                    
                    if line_upper_stripped in alternate_state_name_map:
                        current_state_name = alternate_state_name_map[line_upper_stripped]
                        current_constituency_id = None
                        continue

                    const_match = normal_const_regex.search(line)
                    is_reversed = False
                    if not const_match:
                        const_match = reverse_const_regex.search(line)
                        is_reversed = True

                    if const_match:
                        const_name = ""; cat = "GENERAL"
                        try:
                            if is_reversed:
                                const_name = normalize_name(const_match.group(1))
                                cat_match = re.search(r"\((ST|SC)\)", line, re.I)
                                cat = cat_match.group(1).upper() if cat_match else "GENERAL"
                            else:
                                const_name = normalize_name(const_match.group(2))
                                cat = const_match.group(3).upper() if const_match.group(3) else "GENERAL"
                        except Exception: continue

                        if not const_name or not current_state_name: continue

                        total_electors = 0
                        for j in range(1, 4):
                            if i + j < len(lines):
                                m = re.search(r"\(Total Electors\s*([\d,]+)\)", lines[i + j], re.I)
                                if m: total_electors = safe_int(m.group(1)); break
                        
                        found_id = None
                        state_upper = current_state_name.upper(); const_upper = const_name.upper()

                        if state_upper in state_to_const_map and const_upper in state_to_const_map[state_upper]:
                            found_id = state_to_const_map[state_upper][const_upper]
                        
                        if not found_id: continue

                        current_constituency_id = found_id
                        current_total_electors = total_electors
                        
                        if found_id not in candidates_by_constituency:
                            # Store candidates list inside a dictionary for 2009 for easier handling in merge
                            candidates_by_constituency[found_id] = {"Candidates": [], "Category": cat}
                        continue

                    if not current_constituency_id: continue
                    anchor_match = anchor_regex.search(line)
                    if not anchor_match: continue

                    try:
                        sex = anchor_match.group(1).strip().upper()
                        age = safe_int(anchor_match.group(2))
                        category = anchor_match.group(3).strip().upper()
                        before = line[:anchor_match.start()]
                        after = line[anchor_match.end():]

                        before_match = re.match(r"^\s*(\d+)\s+(.+?)\s*$", before)
                        if not before_match: continue

                        candidate = before_match.group(2).strip()
                        after_match = re.match(r"^\s*(.+?)\s+([\d-]+)\s+([\d-]+)\s+([\d-]+)\s+([\d\.-]+)\s+([\d\.-]+)\s*$", after)
                        if not after_match: continue

                        party = after_match.group(1).strip()
                        gen_votes = safe_int(after_match.group(2))
                        post_votes = safe_int(after_match.group(3))
                        total_votes = safe_int(after_match.group(4))
                        pct_electors = safe_float(after_match.group(5))
                        pct_polled = safe_float(after_match.group(6))

                        candidate_data = {
                            "Candidate Name": candidate, "Gender": "MALE" if sex == "M" else "FEMALE",
                            "Age": age, "Category": category, "Party Name": party, "Party Symbol": None,
                            "Total Votes Polled In The Constituency": 0, "Valid Votes": 0,
                            "Votes Secured": {"General": gen_votes, "Postal": post_votes, "Total": total_votes},
                            "% of Votes Secured": {"Over Total Electors In Constituency": round(pct_electors, 2),
                                                   "Over Total Votes Polled In Constituency": round(pct_polled, 2)},
                            "Over Total Valid Votes Polled In Constituency": 0.0,
                            "Total Electors": current_total_electors
                        }
                        candidates_by_constituency[current_constituency_id]["Candidates"].append(candidate_data)

                    except Exception as e:
                        pass # Silently skip malformed candidate lines

    except Exception as e:
        print(f"Error opening/parsing 2009 detailed PDF: {e}")
        return {}

    print(f"--- 2009 Detailed PDF parsing complete. Found data for {len(candidates_by_constituency)} constituencies. ---")
    return candidates_by_constituency


# --------------------------------------------------------------------------
# 2014 XLSX PARSERS
# --------------------------------------------------------------------------
def parse_2014_summary_sheet(sheet):
    data = {
        "ID": sheet.title.replace(u'\xa0', ' ').strip(), "Constituency": None, "State_UT": None, "Category": None, "Candidates": [], 
        "Summary_Candidate_Stats": {}, "Electors": get_2024_electors_template(), "Voters": get_2024_voters_template(), 
        "Votes": get_2024_votes_template(), "Polling_Station": {"Number": 0, "Average Electors Per Polling": 0},
        "Dates": [], "Result": {"Winner": {"Party": None, "Candidates": None, "Votes": 0}, "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0}, "Margin": 0}
    }
    try:
        state_raw = sheet['B2'].value
        const_raw = str(sheet['D2'].value).replace(u'\xa0', ' ').strip()
        if state_raw: data["State_UT"] = str(state_raw).split('-')[0].replace(u'\xa0', ' ').strip()
        if const_raw:
            cat_match = re.search(r"\((SC|ST)\)", const_raw, re.I)
            data["Category"] = cat_match.group(1).upper() if cat_match else "GENERAL"
            const_name_cleaned = re.sub(r"\s*\((SC|ST)\)\s*$", "", const_raw, flags=re.I)
            const_name_cleaned = re.sub(r"\s*-\s*\d+\s*$", "", const_name_cleaned)
            data["Constituency"] = format_constituency_name(const_name_cleaned)

        data["Summary_Candidate_Stats"]["Contested"] = {"Men": safe_int(sheet['D7'].value), "Women": safe_int(sheet['E7'].value), "Third_Gender": safe_int(sheet['F7'].value), "Total": safe_int(sheet['G7'].value)}
        data["Electors"]["General"] = {"Men": safe_int(sheet['D10'].value), "Women": safe_int(sheet['E10'].value), "Third_Gender": safe_int(sheet['F10'].value), "Total": safe_int(sheet['G10'].value)}
        data["Electors"]["OverSeas"] = {"Men": safe_int(sheet['D11'].value), "Women": safe_int(sheet['E11'].value), "Third_Gender": safe_int(sheet['F11'].value), "Total": safe_int(sheet['G11'].value)}
        data["Electors"]["Service"] = {"Men": safe_int(sheet['D12'].value), "Women": safe_int(sheet['E12'].value), "Third_Gender": safe_int(sheet['F12'].value), "Total": safe_int(sheet['G12'].value)}
        data["Electors"]["Total"] = {"Men": safe_int(sheet['D13'].value), "Women": safe_int(sheet['E13'].value), "Third_Gender": safe_int(sheet['F13'].value), "Total": safe_int(sheet['G13'].value)}
        data["Voters"]["General"] = {"Men": safe_int(sheet['D15'].value), "Women": safe_int(sheet['E15'].value), "Third_Gender": safe_int(sheet['F15'].value), "Total": safe_int(sheet['G15'].value)}
        data["Voters"]["OverSeas"] = {"Men": safe_int(sheet['D16'].value), "Women": safe_int(sheet['E16'].value), "Third_Gender": safe_int(sheet['F16'].value), "Total": safe_int(sheet['G16'].value)}
        data["Voters"]["Proxy"]["Total"] = safe_int(sheet['G17'].value)
        data["Voters"]["Postal"]["Total"] = safe_int(sheet['G18'].value)
        data["Voters"]["Total"]["Total"] = safe_int(sheet['G19'].value)
        data["Voters"]["POLLING PERCENTAGE"]["Total"] = safe_float(sheet['G21'].value)
        data["Votes"]["Total Votes Polled On EVM"] = safe_int(sheet['G23'].value)
        data["Votes"]["Total Deducted Votes From EVM"] = safe_int(sheet['G24'].value)
        data["Votes"]["Total Valid Votes polled on EVM"] = safe_int(sheet['G25'].value)
        data["Votes"]["Postal Votes Counted"] = safe_int(sheet['G26'].value)
        data["Votes"]["Postal Votes Deducted"] = safe_int(sheet['G27'].value)
        data["Votes"]["Valid Postal Votes"] = safe_int(sheet['G28'].value)
        data["Votes"]["Total Valid Votes Polled"] = safe_int(sheet['G29'].value)
        data["Votes"]["Votes Polled for 'NOTA'(Including Postal)"] = safe_int(sheet['G30'].value)
        data["Votes"]["Tendered Votes"] = safe_int(sheet['G31'].value)
        data["Polling_Station"]["Number"] = safe_int(sheet['D33'].value)
        data["Polling_Station"]["Average Electors Per Polling"] = safe_int(sheet['G33'].value)
        polling_date = sheet['D37'].value
        declaration_date = sheet['F37'].value
        if polling_date and not str(polling_date).strip().lower() == "polling": data["Dates"].append(str(polling_date))
        if declaration_date and not str(declaration_date).strip().lower() == "declaration of result": data["Dates"].append(str(declaration_date))
        data["Result"]["Winner"] = {"Party": str(sheet['D39'].value), "Candidates": str(sheet['E39'].value), "Votes": safe_int(sheet['G39'].value)}
        data["Result"]["Runner-Up"] = {"Party": str(sheet['D40'].value), "Candidates": str(sheet['E40'].value), "Votes": safe_int(sheet['G40'].value)}
        data["Result"]["Margin"] = safe_int(sheet['D41'].value)

    except Exception as e:
        print(f"Error parsing 2014 summary sheet {sheet.title}: {e}")
        return None
    return data

def parse_2014_detailed_sheet(sheet, ids):
    candidates_by_constituency = {}
    state_to_const_map = {}
    for full_id, details in ids.items():
        state_upper = details["State_UT"].upper().strip()
        const_upper = details["Constituency"].upper().strip()
        if state_upper not in state_to_const_map: state_to_const_map[state_upper] = {}
        state_to_const_map[state_upper][const_upper] = full_id
    
    alternate_state_name_map = {}
    for name in ids.values():
        state_name = name["State_UT"].upper().strip()
        alternate_state_name_map[state_name] = state_name
    alternate_state_name_map["ORISSA"] = "ODISHA"; alternate_state_name_map["DELHI"] = "NCT OF DELHI"; alternate_state_name_map["NATIONAL CAPITAL TERRITORY OF DELHI"] = "NCT OF DELHI"; alternate_state_name_map["CHATTISGARH"] = "CHHATTISGARH"; alternate_state_name_map["CHHATISGARH"] = "CHHATTISGARH"

    current_state = None
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
        try:
            if not row[2] or str(row[0]).strip().lower() == "state name" or str(row[0]).strip().lower() == "total": continue

            if row[0] and not str(row[0]).strip().startswith('='):
                raw_state_name = str(row[0]).strip().upper().replace(u'\xa0', ' ')
                current_state = alternate_state_name_map.get(raw_state_name, raw_state_name)

            if not current_state: continue

            state = current_state 
            constituency_name_raw = row[1]
            if not constituency_name_raw: continue
            constituency = format_constituency_name(constituency_name_raw)
            
            constituency_id = None
            state_upper = state.upper(); const_upper = constituency.upper()
            if state_upper in state_to_const_map and const_upper in state_to_const_map[state_upper]:
                constituency_id = state_to_const_map[state_upper][const_upper]
            if not constituency_id: continue

            candidate_data = {
                "Candidate Name": str(row[2]).strip(), "Gender": "MALE" if str(row[3]).upper() == "M" else "FEMALE",
                "Age": safe_int(row[4]), "Category": str(row[5]).upper(), "Party Name": str(row[6]).strip(),
                "Party Symbol": str(row[7]).strip(), "Total Votes Polled In The Constituency": 0, "Valid Votes": 0,
                "Votes Secured": {"General": safe_int(row[8]), "Postal": safe_int(row[9]), "Total": safe_int(row[10])},
                "% of Votes Secured": {"Over Total Electors In Constituency": round(safe_float(row[11]), 2), "Over Total Votes Polled In Constituency": round(safe_float(row[12]), 2)},
                "Over Total Valid Votes Polled In Constituency": 0.0, "Total Electors": safe_int(row[13])
            }

            if constituency_id not in candidates_by_constituency: candidates_by_constituency[constituency_id] = []
            candidates_by_constituency[constituency_id].append(candidate_data)
            
        except Exception as e:
            print(f"Error processing 2014 detailed row {row_idx}: {e}")
            
    return candidates_by_constituency


# --------------------------------------------------------------------------
# 2019/2024 XLSX PARSERS
# --------------------------------------------------------------------------
def parse_2019_2024_summary_sheet(sheet, year):
    all_rows = list(sheet.iter_rows(values_only=True))
    data = {
        "ID": sheet.title.strip(), "Constituency": None, "State_UT": None, "Category": None,
        "Candidates": [], "Summary_Candidate_Stats": {}, "Electors": get_2024_electors_template(),
        "Voters": get_2024_voters_template(), "Votes": get_2024_votes_template(),
        "Polling_Station": {"Number": 0, "Average Electors Per Polling": 0}, "Dates": [],
        "Result": {"Winner": {"Party": None, "Candidates": None, "Votes": 0}, "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0}, "Margin": 0}
    }
    
    current_section = CurrentSection.NONE
    for i, row in enumerate(all_rows):
        if not any(row): continue 
        cell_one = str(clean_value(row[0]))
        
        if "State/UT" in cell_one:
            current_section = CurrentSection.STATE_UT
            state_name = str(clean_value(row[1])).split('-')[0].strip()
            data["State_UT"] = STATE_NAME_CORRECTIONS.get(state_name.lower(), state_name)
            const_raw = str(clean_value(row[3])).strip()
            cat_match = re.search(r"\((SC|ST)\)", const_raw, re.I); 
            if not cat_match: cat_match = re.search(r"-(SC|ST)", const_raw, re.I)
            data["Category"] = cat_match.group(1).upper() if cat_match else "GENERAL"
            data["Constituency"] = format_constituency_name(const_raw)
            
        elif "CANDIDATES" in cell_one: current_section = CurrentSection.SUMMARY_CANDIDATE_STATS
        elif "ELECTORS" in cell_one: current_section = CurrentSection.ELECTORS
        elif "VOTERS" in cell_one: current_section = CurrentSection.VOTERS
        elif "VOTES" in cell_one: current_section = CurrentSection.VOTES
        elif "POLLING STATION" in cell_one: current_section = CurrentSection.POLLING_STATION
        elif "DATES" in cell_one:
            current_section = CurrentSection.DATES
            date_row = None
            for j in range(i, min(i + 5, len(all_rows))):
                test_row = all_rows[j]
                val3 = clean_value(test_row[3]); val5 = clean_value(test_row[5])
                is_valid_date_content = (isinstance(val3, str) and '/' in val3) or (isinstance(val5, str) and '/' in val5)
                if is_valid_date_content: date_row = test_row; break
            
            if date_row:
                poll_date = str(clean_value(date_row[3])); decl_date = str(clean_value(date_row[5]))
                if poll_date and "/" in poll_date and "polling" not in poll_date.lower(): data["Dates"].append(poll_date)
                poll_date_alt = str(clean_value(date_row[4]))
                if poll_date_alt and poll_date_alt not in data["Dates"] and "/" in poll_date_alt and "polling" not in poll_date_alt.lower(): data["Dates"].append(poll_date_alt)
                if decl_date and "/" in decl_date and "declaration" not in decl_date.lower(): data["Dates"].append(decl_date)
            
        elif "RESULT" in cell_one: current_section = CurrentSection.RESULT
            
        elif current_section == CurrentSection.SUMMARY_CANDIDATE_STATS or current_section == CurrentSection.ELECTORS:
            key = str(clean_value(row[1]))
            if key and len(row) > 6: data[current_section.value][key] = {"Men": safe_int(row[3]), "Women": safe_int(row[4]), "Third_Gender": safe_int(row[5]), "Total": safe_int(row[6])}
        elif current_section == CurrentSection.VOTERS:
            key = str(clean_value(row[1]))
            if not key: continue
            if "POLLING PERCENTAGE" in key:
                pct_val = row[3] if safe_float(row[3]) != 0.0 else row[6]
                data["Voters"]["POLLING PERCENTAGE"]["Total"] = safe_float(pct_val)
            elif key in data[current_section.value] and len(row) > 6: data[current_section.value][key] = {"Men": safe_int(row[3]), "Women": safe_int(row[4]), "Third_Gender": safe_int(row[5]), "Total": safe_int(row[6])}
            
        elif current_section == CurrentSection.VOTES:
            key = str(clean_value(row[1]))
            if key and len(row) > 6 and key in data[current_section.value]: data[current_section.value][key] = safe_int(row[6])
            
        elif current_section == CurrentSection.POLLING_STATION:
            key = str(clean_value(row[1]))
            if key == "Number": data[current_section.value][key] = safe_int(row[3])
            elif "Average Electors" in key: data[current_section.value]["Average Electors Per Polling"] = safe_int(row[6]); current_section = CurrentSection.NONE 
                 
        elif current_section == CurrentSection.RESULT and not data["Result"]["Winner"]["Party"]:
            key = str(clean_value(row[1]))
            if key in ["Winner", "Runner-Up"] and len(row) > 6: data[current_section.value][key] = {"Party": clean_value(row[3]), "Candidates": clean_value(row[4]), "Votes": safe_int(row[6])}
            elif key == "Margin": data[current_section.value][key] = safe_int(row[3]); current_section = CurrentSection.NONE

    data["Dates"] = sorted(list(set(data["Dates"])))
    return data

def parse_2019_2024_detailed_sheet(sheet, ids, year, header_map):
    candidates = defaultdict(list)
    rows = list(sheet.iter_rows(values_only=True))
    header_row_index = 1; subheader_row_index = 2; data_start_row = 3
    if year <= 2014: header_row_index = 0; subheader_row_index = 1; data_start_row = 2
        
    l1_fields = [str(clean_value(h)).replace('\n', ' ').strip().lower() if h else "" for h in rows[header_row_index]]
    l2_fields = [str(clean_value(h)).replace('\n', ' ').strip().lower() if h else "" for h in rows[subheader_row_index]]

    last_valid_header = ""
    for i in range(len(l1_fields)):
        if l1_fields[i]: last_valid_header = l1_fields[i]
        else: l1_fields[i] = last_valid_header

    constituency_lookup = {(v['State_UT'].lower(), v['Constituency'].lower()): k for k, v in ids.items() if v['State_UT'] and v['Constituency']}
    header_map["% Over Total Valid Votes"] = -1
    for i in range(len(l2_fields)):
        l1 = l1_fields[i]; l2 = l2_fields[i]
        if "candidate" in l2 and "name" in l2: header_map["Candidate Name"] = i
        elif l2 == "sex" or l2 == "gender": header_map["Gender"] = i 
        elif l2 == "age": header_map["Age"] = i
        elif l2 == "category": header_map["Category"] = i
        elif "party name" in l2: header_map["Party Name"] = i
        elif "party symbol" in l2: header_map["Party Symbol"] = i
        elif l2 == "total votes polled in the constituency": header_map["Total Votes Polled"] = i
        elif l2 == "valid votes": header_map["Valid Votes"] = i
        elif "votes secured" in l1 and l2 == "general": header_map["General"] = i
        elif "votes secured" in l1 and l2 == "postal": header_map["Postal"] = i
        elif "votes secured" in l1 and l2 == "total": header_map["Total"] = i
        elif "% of votes secured" in l1 and "over total electors" in l2: header_map["% Over Total Electors"] = i
        elif "% of votes secured" in l1 and "over total votes polled" in l2: header_map["% Over Total Votes Polled"] = i
        elif "% of votes secured" in l1 and "over total valid votes" in l2: header_map["% Over Total Valid Votes"] = i
        elif "total electors" in l1 or "total electors" in l2: header_map["Total Electors"] = i

    if header_map["% Over Total Valid Votes"] == -1: header_map["% Over Total Valid Votes"] = -2
        
    for row in rows[data_start_row:]:
        if header_map["Candidate Name"] == -1 or not row[header_map["Candidate Name"]]: continue
        state = clean_value(row[0]); state_standardized = STATE_NAME_CORRECTIONS.get(state.lower(), state)
        constituency = format_constituency_name(clean_value(row[1])); lookup_key = (state_standardized.lower(), constituency.lower())
        constituency_id = constituency_lookup.get(lookup_key)
        
        if not constituency_id: continue

        try:
            pct_valid_votes = 0.0
            if header_map["% Over Total Valid Votes"] >= 0: pct_valid_votes = round(safe_float(row[header_map["% Over Total Valid Votes"]]), 2)
            
            candidate_data = {
                "Candidate Name": clean_value(row[header_map.get("Candidate Name", -1)]), "Gender": clean_value(row[header_map.get("Gender", -1)]),
                "Age": safe_int(row[header_map.get("Age", -1)]), "Category": clean_value(row[header_map.get("Category", -1)]),
                "Party Name": clean_value(row[header_map.get("Party Name", -1)]), "Party Symbol": clean_value(row[header_map.get("Party Symbol", -1)]),
                "Total Votes Polled In The Constituency": safe_int(row[header_map.get("Total Votes Polled", -1)]), "Valid Votes": safe_int(row[header_map.get("Valid Votes", -1)]),
                "Votes Secured": {"General": safe_int(row[header_map.get("General", -1)]), "Postal": safe_int(row[header_map.get("Postal", -1)]), "Total": safe_int(row[header_map.get("Total", -1)])},
                "% of Votes Secured": {"Over Total Electors In Constituency": round(safe_float(row[header_map.get("% Over Total Electors", -1)]), 2), "Over Total Votes Polled In Constituency": round(safe_float(row[header_map.get("% Over Total Votes Polled", -1)]), 2)},
                "Over Total Valid Votes Polled In Constituency": pct_valid_votes, "Total Electors": safe_int(row[header_map.get("Total Electors", -1)])
            }
            candidates[constituency_id].append(candidate_data)
        except Exception as e:
            pass
            
    return candidates


# --------------------------------------------------------------------------
# MAIN EXECUTION AND MERGE LOGIC
# --------------------------------------------------------------------------

def parse_and_merge(year, summary_path, detailed_path, output_path, parser_type):
    print(f"\n--- Starting processing for year: *{year}* ({parser_type}) ---")
    print(f"  Summary: {summary_path}")
    print(f"  Detailed: {detailed_path}")
    
    parsed_summary = []
    candidates_map = {}

    try:
        if parser_type == "PDF":
            parsed_summary = parse_2009_summary_pdf(summary_path)
            ids = {c["ID"]: {"State_UT": c["State_UT"], "Constituency": c["Constituency"]} for c in parsed_summary if c["ID"]}
            candidates_map = parse_2009_detailed_pdf(detailed_path, ids)
            
        elif parser_type == "XLSX":
            if 'openpyxl' not in sys.modules: print("Error: 'openpyxl' not installed. Skipping XLSX parsing."); return
            wb_summary = load_workbook(summary_path, data_only=True)
            if year == 2014:
                parsed_summary = [parse_2014_summary_sheet(wb_summary[s]) for s in wb_summary.sheetnames]; parsed_summary = [p for p in parsed_summary if p]
            else: parsed_summary = [parse_2019_2024_summary_sheet(wb_summary[s], year) for s in wb_summary.sheetnames]
            
            print(f"Parsed {len(parsed_summary)} constituency summaries.")
            ids = {c["ID"]: {"State_UT": c["State_UT"], "Constituency": c["Constituency"]} for c in parsed_summary if c["ID"]}
            wb_detailed = load_workbook(detailed_path, data_only=True); active_sheet = wb_detailed.active
            
            if year == 2014: candidates_map = parse_2014_detailed_sheet(active_sheet, ids)
            else: candidates_map = parse_2019_2024_detailed_sheet(active_sheet, ids, year, defaultdict(lambda: -1))

        print(f"Parsed candidate data for {len(candidates_map)} constituencies.")
        
        merged_count = 0
        for constituency_summary in parsed_summary:
            full_id = constituency_summary.get('ID')
            if not full_id or full_id not in candidates_map: 
                if 'Summary_Candidate_Stats' in constituency_summary: del constituency_summary['Summary_Candidate_Stats']
                continue

            if parser_type == "PDF":
                candidate_data_container = candidates_map[full_id]
                candidate_list = candidate_data_container.get("Candidates", [])
            else:
                candidate_list = candidates_map[full_id]

            total_polled = constituency_summary.get("Voters", {}).get("Total", {}).get("Total", 0)
            valid_votes_from_summary = constituency_summary.get("Votes", {}).get("Total Valid Votes Polled", 0)
            
            # Update Candidate List and Stats
            for cand in candidate_list:
                if cand["Total Votes Polled In The Constituency"] == 0:
                    cand["Total Votes Polled In The Constituency"] = total_polled
                
                valid_votes_to_use = cand["Valid Votes"] if cand["Valid Votes"] != 0 else valid_votes_from_summary
                if cand["Valid Votes"] == 0: cand["Valid Votes"] = valid_votes_from_summary
                
                if valid_votes_to_use > 0:
                    cand["Over Total Valid Votes Polled In Constituency"] = round(
                        (cand["Votes Secured"]["Total"] / valid_votes_to_use) * 100, 2
                    )
                else:
                    cand["Over Total Valid Votes Polled In Constituency"] = 0.0

            constituency_summary['Candidates'] = candidate_list
            
            # Re-calculate Winner/Runner-Up from merged list
            if candidate_list:
                sorted_candidates = sorted(candidate_list, key=lambda c: c["Votes Secured"]["Total"], reverse=True)
                if sorted_candidates:
                    winner = sorted_candidates[0]
                    constituency_summary["Result"]["Winner"] = {"Party": winner["Party Name"], "Candidates": winner["Candidate Name"], "Votes": winner["Votes Secured"]["Total"]}
                    
                    if len(sorted_candidates) > 1:
                        runner_up = sorted_candidates[1]
                        constituency_summary["Result"]["Runner-Up"] = {"Party": runner_up["Party Name"], "Candidates": runner_up["Candidate Name"], "Votes": runner_up["Votes Secured"]["Total"]}
                        constituency_summary["Result"]["Margin"] = winner["Votes Secured"]["Total"] - runner_up["Votes Secured"]["Total"]
                    elif len(sorted_candidates) > 0:
                        constituency_summary["Result"]["Runner-Up"] = {"Party": None, "Candidates": None, "Votes": 0}
                        constituency_summary["Result"]["Margin"] = winner["Votes Secured"]["Total"]
            
            merged_count += 1
            if 'Summary_Candidate_Stats' in constituency_summary: del constituency_summary['Summary_Candidate_Stats']

        print(f"Merged detailed candidate data for {merged_count} constituencies.")

        # 4. Dump to JSON
        output_dir = os.path.dirname(output_path)
        if output_dir: os.makedirs(output_dir, exist_ok=True)

        print(f"Writing final JSON to: {output_path}")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(parsed_summary, f, indent=4, default=str)
        print("JSON file written successfully.")

    except FileNotFoundError as e:
        print(f"\nFATAL ERROR (Year {year}): File not found at path: {e.filename}")
    except Exception as e:
        print(f"\nAn unexpected error occurred during Year {year} parsing: {e}")
        import traceback
        traceback.print_exc()

def run_all_parsers():
    for year, config in JOBS_CONFIG.items():
        parse_and_merge(
            year, 
            config["summary_path"], 
            config["detailed_path"], 
            config["output_path"], 
            config["parser_type"]
        )

# --------------------------------------------------------------------------
if __name__ == "__main__":
    run_all_parsers()