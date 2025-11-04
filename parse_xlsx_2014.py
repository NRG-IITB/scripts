"""
Script to parse 2014 constituency-wise data reports (XLSX format).
This script uses openpyxl to read .xlsx files (ECI .xls files must be
pre-converted to .xlsx) and produces JSON output matching the 2024
JSON format.
"""

from openpyxl import load_workbook
import json
import string
import sys
import re

# --------------------------------------------------------------------------
# HELPERS (from 2009 script)
# --------------------------------------------------------------------------

def format_constituency_name(name):
    """
    Cleans and formats constituency names.
    Removes (SC)/(ST) and title-cases the name.
    """
    if not name:
        return ""
    
    # Replace non-breaking spaces with regular spaces
    name = str(name).replace(u'\xa0', ' ')

    # Remove leading number/hyphen, e.g., "1-"
    name = re.sub(r"^\s*[\d\s-]+\s*", "", name)
    
    # Remove (SC) or (ST) suffixes
    name = re.sub(r"\s*\((SC|ST)\)\s*$", "", name, flags=re.I)
    
    # Handle names like "ADILABAD -1"
    name = re.sub(r"\s*-\s*\d+\s*$", "", name)

    parts = name.split('-')
    parts = [string.capwords(part.strip()) for part in parts]
    return '-'.join(parts).replace('&', 'and')

def safe_int(value):
    """Safely convert value to integer, handling None, formulas, and commas."""
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '')
        if value.startswith('(') and value.endswith(')'): # Handle "(0)"
            value = value[1:-1]
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0

def safe_float(value):
    """Safely convert value to float, handling None, formulas, and commas."""
    if isinstance(value, (float, int)):
        return float(value)
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '')
        if value.startswith('(') and value.endswith(')'): # Handle "(0)"
            value = value[1:-1]
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

# --------------------------------------------------------------------------
# SUMMARY XLSX PARSER (Report 32)
# --------------------------------------------------------------------------

def parse_summary_sheet(sheet):
    """
    Parses a single constituency sheet from the 2014 summary workbook.
    This version reads from fixed cell locations for reliability.
    """
    
    # Create the target 2024 JSON structure
    data = {
        "ID": sheet.title.replace(u'\xa0', ' ').strip(), # Use sheet name as ID
        "Constituency": None,
        "State_UT": None,
        "Category": None,
        "Candidates": [], # To be filled by detailed parser
        "Summary_Candidate_Stats": {}, # This will be removed at the end to match 2024
        "Electors": {
            "General": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0},
            "OverSeas": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0}, # 2014+
            "Service": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0},
            "Total": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0},
        },
        "Voters": {
            "General": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0},
            "OverSeas": {"Men": 0, "Women": 0, "Third_Gender": 0, "Total": 0}, # 2014+
            "Proxy": {"Men": None, "Women": None, "Third_Gender": None, "Total": 0},
            "Postal": {"Men": None, "Women": None, "Third_Gender": None, "Total": 0},
            "Total": {"Men": None, "Women": None, "Third_Gender": None, "Total": 0},
            "Votes Not Counted From CU(s) as Per ECI Instructions": {
                "Men": None, "Women": None, "Third_Gender": None, "Total": 0
            },
            "POLLING PERCENTAGE": {
                 "Men": None, "Women": None, "Third_Gender": None, "Total": 0.0
            },
        },
        "Votes": {
            "Total Votes Polled On EVM": 0,
            "Total Deducted Votes From EVM": 0,
            "Total Valid Votes polled on EVM": 0,
            "Postal Votes Counted": 0,
            "Postal Votes Deducted": 0,
            "Valid Postal Votes": 0,
            "Total Valid Votes Polled": 0,
            "Test Votes polled On EVM": 0,
            "Votes Polled for 'NOTA'(Including Postal)": 0, # 2014+
            "Tendered Votes": 0,
        },
        # --- [SCHEMA FIX] ---
        # Cleaned Polling_Station object to match 2024 standard
        "Polling_Station": {
            "Number": 0,
            "Average Electors Per Polling": 0,
        },
        # --- [END SCHEMA FIX] ---
        "Dates": [],
        "Result": {
            "Winner": {"Party": None, "Candidates": None, "Votes": 0},
            "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0},
            "Margin": 0,
        }
    }
    
    try:
        # --- Header ---
        state_raw = sheet['B2'].value
        const_raw = str(sheet['D2'].value).replace(u'\xa0', ' ').strip()
        
        if state_raw:
            data["State_UT"] = state_raw.split('-')[0].replace(u'\xa0', ' ').strip()
        
        if const_raw:
            cat_match = re.search(r"\((SC|ST)\)", const_raw, re.I)
            if cat_match:
                data["Category"] = cat_match.group(1).upper()
            else:
                data["Category"] = "GENERAL"
            
            const_name_cleaned = re.sub(r"\s*\((SC|ST)\)\s*$", "", const_raw, flags=re.I)
            const_name_cleaned = re.sub(r"\s*-\s*\d+\s*$", "", const_name_cleaned)
            data["Constituency"] = format_constituency_name(const_name_cleaned)

        # --- I. CANDIDATES (Summary_Candidate_Stats) ---
        data["Summary_Candidate_Stats"]["Nominated"] = {"Men": safe_int(sheet['D4'].value), "Women": safe_int(sheet['E4'].value), "Third_Gender": safe_int(sheet['F4'].value), "Total": safe_int(sheet['G4'].value)}
        data["Summary_Candidate_Stats"]["Nomination Rejected"] = {"Men": safe_int(sheet['D5'].value), "Women": safe_int(sheet['E5'].value), "Third_Gender": safe_int(sheet['F5'].value), "Total": safe_int(sheet['G5'].value)}
        data["Summary_Candidate_Stats"]["Withdrawn"] = {"Men": safe_int(sheet['D6'].value), "Women": safe_int(sheet['E6'].value), "Third_Gender": safe_int(sheet['F6'].value), "Total": safe_int(sheet['G6'].value)}
        data["Summary_Candidate_Stats"]["Contested"] = {"Men": safe_int(sheet['D7'].value), "Women": safe_int(sheet['E7'].value), "Third_Gender": safe_int(sheet['F7'].value), "Total": safe_int(sheet['G7'].value)}
        data["Summary_Candidate_Stats"]["Forfeited Deposit"] = {"Men": safe_int(sheet['D8'].value), "Women": safe_int(sheet['E8'].value), "Third_Gender": safe_int(sheet['F8'].value), "Total": safe_int(sheet['G8'].value)}
        
        # --- II. ELECTORS ---
        data["Electors"]["General"] = {"Men": safe_int(sheet['D10'].value), "Women": safe_int(sheet['E10'].value), "Third_Gender": safe_int(sheet['F10'].value), "Total": safe_int(sheet['G10'].value)}
        data["Electors"]["OverSeas"] = {"Men": safe_int(sheet['D11'].value), "Women": safe_int(sheet['E11'].value), "Third_Gender": safe_int(sheet['F11'].value), "Total": safe_int(sheet['G11'].value)}
        data["Electors"]["Service"] = {"Men": safe_int(sheet['D12'].value), "Women": safe_int(sheet['E12'].value), "Third_Gender": safe_int(sheet['F12'].value), "Total": safe_int(sheet['G12'].value)}
        data["Electors"]["Total"] = {"Men": safe_int(sheet['D13'].value), "Women": safe_int(sheet['E13'].value), "Third_Gender": safe_int(sheet['F13'].value), "Total": safe_int(sheet['G13'].value)}
        
        # --- III. VOTERS ---
        data["Voters"]["General"] = {"Men": safe_int(sheet['D15'].value), "Women": safe_int(sheet['E15'].value), "Third_Gender": safe_int(sheet['F15'].value), "Total": safe_int(sheet['G15'].value)}
        data["Voters"]["OverSeas"] = {"Men": safe_int(sheet['D16'].value), "Women": safe_int(sheet['E16'].value), "Third_Gender": safe_int(sheet['F16'].value), "Total": safe_int(sheet['G16'].value)}
        data["Voters"]["Proxy"]["Total"] = safe_int(sheet['G17'].value)
        data["Voters"]["Postal"]["Total"] = safe_int(sheet['G18'].value) # G18 or I18
        data["Voters"]["Total"]["Total"] = safe_int(sheet['G19'].value)
        data["Voters"]["POLLING PERCENTAGE"]["Total"] = safe_float(sheet['G21'].value) # G21
        
        # --- IV. VOTES ---
        data["Votes"]["Total Votes Polled On EVM"] = safe_int(sheet['G23'].value)
        data["Votes"]["Total Deducted Votes From EVM"] = safe_int(sheet['G24'].value)
        data["Votes"]["Total Valid Votes polled on EVM"] = safe_int(sheet['G25'].value)
        data["Votes"]["Postal Votes Counted"] = safe_int(sheet['G26'].value)
        data["Votes"]["Postal Votes Deducted"] = safe_int(sheet['G27'].value)
        data["Votes"]["Valid Postal Votes"] = safe_int(sheet['G28'].value)
        data["Votes"]["Total Valid Votes Polled"] = safe_int(sheet['G29'].value)
        data["Votes"]["Test Votes polled On EVM"] = 0 # Not present
        data["Votes"]["Votes Polled for 'NOTA'(Including Postal)"] = safe_int(sheet['G30'].value)
        data["Votes"]["Tendered Votes"] = safe_int(sheet['G31'].value)
        
        # --- V. POLLING STATION (SCHEMA FIX) ---
        # Match 2024 schema
        data["Polling_Station"]["Number"] = safe_int(sheet['D33'].value)
        data["Polling_Station"]["Average Electors Per Polling"] = safe_int(sheet['G33'].value)
        # Removed logic for Re-Poll Dates and Numbers to match 2024
        
        # --- VI. DATES (SCHEMA FIX) ---
        # Match 2024 format: [Polling Date, Declaration Date]
        polling_date = sheet['D37'].value
        declaration_date = sheet['F37'].value
        
        # Add checks to avoid saving placeholder text
        if polling_date and not str(polling_date).strip().lower() == "polling":
            data["Dates"].append(str(polling_date))
        if declaration_date and not str(declaration_date).strip().lower() == "declaration of result":
            data["Dates"].append(str(declaration_date))
        # --- [END SCHEMA FIX] ---

        # --- VII. RESULT ---
        # This data is unreliable and will be RECALCULATED after merging
        data["Result"]["Winner"] = {"Party": str(sheet['D39'].value), "Candidates": str(sheet['E39'].value), "Votes": safe_int(sheet['G39'].value)}
        data["Result"]["Runner-Up"] = {"Party": str(sheet['D40'].value), "Candidates": str(sheet['E40'].value), "Votes": safe_int(sheet['G40'].value)}
        data["Result"]["Margin"] = safe_int(sheet['D41'].value)

    except Exception as e:
        print(f"Error parsing summary sheet {sheet.title}: {e}")
        print(f"   > Check if file is 2014 format and not corrupt.")
        return None

    return data

def parse_2014_summary_workbook(summary_file_path):
    """
    Loads the 2014 summary workbook and parses every sheet.
    """
    try:
        wb = load_workbook(summary_file_path, data_only=True) # data_only=True to get values, not formulas
        parsed = []
        print(f"Found {len(wb.sheetnames)} sheets in summary file.")
        
        for s in wb.sheetnames:
            sheet = wb[s]
            sheet_data = parse_summary_sheet(sheet)
            if sheet_data:
                parsed.append(sheet_data)
        
        print(f"Successfully parsed {len(parsed)} summary sheets.")
        return parsed
    except Exception as e:
        print(f"Error loading summary workbook: {e}")
        return []

# --------------------------------------------------------------------------
# DETAILED XLSX PARSER (Report 33)
# --------------------------------------------------------------------------

def parse_detailed_sheet(sheet, ids):
    """
    Parses the single detailed results sheet from 2014.
    This version uses fixed column indices based on the 2014 file structure.
    """
    candidates_by_constituency = {}
    
    # --- Create the State-Aware Nested Map ---
    state_to_const_map = {}
    for full_id, details in ids.items():
        state_upper = details["State_UT"].upper().strip()
        const_upper = details["Constituency"].upper().strip()
        if state_upper not in state_to_const_map:
            state_to_const_map[state_upper] = {}
        state_to_const_map[state_upper][const_upper] = full_id
    
    print(f"Built nested state-aware map for {len(state_to_const_map)} states.")

    # --- Create an ALIAS MAP for state names ---
    alternate_state_name_map = {}
    for name in ids.values():
        state_name = name["State_UT"].upper().strip()
        if state_name not in alternate_state_name_map:
            alternate_state_name_map[state_name] = state_name

    # Add aliases (Dirty/Old Name -> Clean/Summary Name)
    alternate_state_name_map["ORISSA"] = "ODISHA" # S18
    alternate_state_name_map["DELHI"] = "NCT OF DELHI" 
    alternate_state_name_map["NATIONAL CAPITAL TERRITORY OF DELHI"] = "NCT OF DELHI" 
    alternate_state_name_map["CHATTISGARH"] = "CHHATTISGARH"
    alternate_state_name_map["CHHATISGARH"] = "CHHATTISGARH"

    current_state = None # This will hold the CANONICAL name
    
    # Iterate from row 3 (data_start_row)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
        try:
            # skip empty, header, and total rows
            if not row[2] or str(row[0]).strip().lower() == "state name" or str(row[0]).strip().lower() == "total":
                continue

            # --- State Tracking (with Alias Map) ---
            if row[0] and not str(row[0]).strip().startswith('='):
                raw_state_name = row[0].strip().upper().replace(u'\xa0', ' ')
                
                if raw_state_name in alternate_state_name_map:
                    current_state = alternate_state_name_map[raw_state_name]
                else:
                    current_state = raw_state_name
                    if current_state not in getattr(parse_detailed_sheet, "warned_state", set()):
                        print(f"Warning: Unknown state '{raw_state_name}' in detailed file. Using it directly.")
                        if not hasattr(parse_detailed_sheet, "warned_state"):
                            parse_detailed_sheet.warned_state = set()
                        parse_detailed_sheet.warned_state.add(current_state)

            if not current_state:
                continue

            state = current_state 
            constituency_name_raw = row[1]
            if not constituency_name_raw:
                continue
                
            constituency = format_constituency_name(constituency_name_raw)
            
            # --- State-Aware ID Matching ---
            constituency_id = None
            state_upper = state.upper()
            const_upper = constituency.upper()
            
            if state_upper in state_to_const_map and const_upper in state_to_const_map[state_upper]:
                constituency_id = state_to_const_map[state_upper][const_upper]
            else:
                if state_upper in state_to_const_map:
                    for known_name, full_id in state_to_const_map[state_upper].items():
                        if const_upper and known_name and len(const_upper) >= 4 and (known_name.startswith(const_upper[:4]) or const_upper.startswith(known_name[:4])):
                            constituency_id = full_id
                            break
            
            if not constituency_id:
                if (state, constituency) not in getattr(parse_detailed_sheet, "warned", set()):
                    print(f"Warning: Could not find ID for {constituency} in {state} (Row {row_idx})")
                    if not hasattr(parse_detailed_sheet, "warned"):
                        parse_detailed_sheet.warned = set()
                    parse_detailed_sheet.warned.add((state, constituency))
                continue

            # --- [SCHEMA/FORMAT FIX] ---
            # Map columns to 2024 JSON structure and round floats
            candidate_data = {
                "Candidate Name": str(row[2]).strip(),
                "Gender": "MALE" if str(row[3]).upper() == "M" else "FEMALE",
                "Age": safe_int(row[4]),
                "Category": str(row[5]).upper(),
                "Party Name": str(row[6]).strip(),
                "Party Symbol": str(row[7]).strip(),
                "Total Votes Polled In The Constituency": 0, # Placeholder
                "Valid Votes": 0, # Placeholder
                "Votes Secured": {
                    "General": safe_int(row[8]),
                    "Postal": safe_int(row[9]),
                    "Total": safe_int(row[10])
                },
                "% of Votes Secured": {
                    "Over Total Electors In Constituency": round(safe_float(row[11]), 2),
                    "Over Total Votes Polled In Constituency": round(safe_float(row[12]), 2),
                },
                # Move keys to top-level of candidate object
                
                "Over Total Valid Votes Polled In Constituency": 0.0, # Placeholder
                "Total Electors": safe_int(row[13])
            }
            # --- [END SCHEMA/FORMAT FIX] ---

            if constituency_id not in candidates_by_constituency:
                candidates_by_constituency[constituency_id] = []
            
            candidates_by_constituency[constituency_id].append(candidate_data)
            
        except Exception as e:
            print(f"Error processing detailed row {row_idx}: {row}")
            print(f"   > Error: {e}")
            
    return candidates_by_constituency


def parse_2014_detailed_workbook(detailed_file_path, ids):
    """
    Loads the 2014 detailed workbook and parses the active sheet.
    """
    try:
        print(f"Loading detailed results file...")
        wb = load_workbook(detailed_file_path, data_only=True)
        active_sheet = wb.active
        print(f"Parsing detailed results sheet: {active_sheet.title}")
        candidates = parse_detailed_sheet(active_sheet, ids)
        print(f"Successfully parsed candidate data for {len(candidates)} constituencies.")
        return candidates
    except Exception as e:
        print(f"Error loading detailed workbook: {e}")
        return {}

# --------------------------------------------------------------------------
# MAIN DRIVER
# --------------------------------------------------------------------------
def run_2014_parse():
    # ** UPDATE THESE PATHS for your 2014 files **
    PDF_PATHS = {
        "summary": "/mnt/d/IIT_B_Project/SL_project/scripts/converted_xlsx_reports/election_data_2014_xlsx/Constituency data summary.xlsx",
        "detailed": "scripts/converted_xlsx_reports/election_data_2014_xlsx/Constituency wise detailed result.xlsx",
        "output": "parsed2014.json",
    }
    
    # Check if file paths were passed as arguments
    if len(sys.argv) == 4:
        PDF_PATHS["summary"] = sys.argv[1]
        PDF_PATHS["detailed"] = sys.argv[2]
        PDF_PATHS["output"] = sys.argv[3]
    
    print(f"Starting 2014 XLSX parsing process...")
    print(f"   Summary file: {PDF_PATHS['summary']}")
    print(f"   Detailed file: {PDF_PATHS['detailed']}")
    print(f"   Output file: {PDF_PATHS['output']}")

    try:
        # 1. Parse the summary workbook
        parsed_summary = parse_2014_summary_workbook(PDF_PATHS["summary"])
        
        if not parsed_summary:
            print("Error: No data parsed from summary workbook. Exiting.")
            return

        # 2. Extract IDs and create a map
        ids = {c["ID"]: {"State_UT": c["State_UT"], "Constituency": c["Constituency"]} for c in parsed_summary if c["ID"]}
        print(f"Extracted {len(ids)} constituency IDs from summary.")

        # 3. Parse the detailed candidate workbook
        candidates_map = parse_2014_detailed_workbook(PDF_PATHS["detailed"], ids)

        # 4. Merge data
        print("\n--- Merging Summary and Detailed Data ---")
        merged_count = 0
        for constituency_summary in parsed_summary:
            full_id = constituency_summary['ID']
            
            if full_id in candidates_map:
                candidate_list = candidates_map[full_id]
                
                # Get total/valid votes from the summary
                total_polled = constituency_summary.get("Voters", {}).get("Total", {}).get("Total", 0)
                valid_votes = constituency_summary.get("Votes", {}).get("Total Valid Votes Polled", 0)
                
                # Update category in summary from detailed (often more accurate)
                if candidate_list:
                     constituency_summary["Category"] = candidate_list[0].get("Category", constituency_summary.get("Category"))

                # Add summary data to each candidate
                for cand in candidate_list:
                    cand["Total Votes Polled In The Constituency"] = total_polled
                    cand["Valid Votes"] = valid_votes
                    
                    # Calculate % over valid votes and add to top-level of candidate
                    if valid_votes > 0:
                        cand["Over Total Valid Votes Polled In Constituency"] = round(
                            (cand["Votes Secured"]["Total"] / valid_votes) * 100, 2
                        )
                    else:
                        cand["Over Total Valid Votes Polled In Constituency"] = 0.0
                
                constituency_summary['Candidates'] = candidate_list
                
                # --- [DATA FIX] ---
                # Re-calculate Winner/Runner-Up from the detailed candidate list
                if candidate_list:
                    try:
                        # Sort candidates by total votes, descending
                        sorted_candidates = sorted(
                            candidate_list, 
                            key=lambda c: c["Votes Secured"]["Total"], 
                            reverse=True
                        )
                        
                        # Update Winner
                        if len(sorted_candidates) > 0:
                            winner = sorted_candidates[0]
                            constituency_summary["Result"]["Winner"] = {
                                "Party": winner["Party Name"],
                                "Candidates": winner["Candidate Name"],
                                "Votes": winner["Votes Secured"]["Total"]
                            }
                        
                        # Update Runner-Up
                        if len(sorted_candidates) > 1:
                            runner_up = sorted_candidates[1]
                            constituency_summary["Result"]["Runner-Up"] = {
                                "Party": runner_up["Party Name"],
                                "Candidates": runner_up["Candidate Name"],
                                "Votes": runner_up["Votes Secured"]["Total"]
                            }
                            
                            # Update Margin
                            constituency_summary["Result"]["Margin"] = (
                                winner["Votes Secured"]["Total"] - runner_up["Votes Secured"]["Total"]
                            )
                        else:
                            # No runner-up, set to None and 0
                            constituency_summary["Result"]["Runner-Up"] = {
                                "Party": None,
                                "Candidates": None,
                                "Votes": 0
                            }
                            constituency_summary["Result"]["Margin"] = winner["Votes Secured"]["Total"]

                    except Exception as e:
                        print(f"Error calculating winner for {full_id}: {e}")
                # --- [END DATA FIX] ---
                
                merged_count += 1
            else:
                print(f"Warning: No detailed candidate data found for {constituency_summary.get('Constituency')} (ID: {full_id})")
                # Remove the empty 'Candidates' key
                if 'Candidates' in constituency_summary:
                    del constituency_summary['Candidates']
            
            # --- [SCHEMA FIX] ---
            # Remove 2014-specific key to match 2024 format
            if 'Summary_Candidate_Stats' in constituency_summary:
                del constituency_summary['Summary_Candidate_Stats']
            # --- [END SCHEMA FIX] ---


        print(f"Merged detailed candidate data for {merged_count} constituencies.")
        
        # Final check
        if merged_count == len(parsed_summary):
             print(f"\nSuccess! Parsed and merged all {merged_count} constituencies.")
        else:
             print(f"\nWarning: Mismatched count! Parsed {len(parsed_summary)} summaries but merged {merged_count} candidate lists.")

        # 5. Dump to JSON
        print(f"\nWriting final merged data to: {PDF_PATHS['output']}")
        with open(PDF_PATHS['output'], 'w') as f:
            json.dump(parsed_summary, f, indent=4)
        print("JSON file written successfully.")

    except FileNotFoundError as e:
        print(f"Error: File not found. Make sure this file exists:")
        print(f"   {e.filename}")
    except Exception as e:
        print(f"An unexpected error occurred during 2014 parsing: {e}")
        import traceback
        traceback.print_exc()

# --------------------------------------------------------------------------
if __name__ == "__main__":
    run_2014_parse()