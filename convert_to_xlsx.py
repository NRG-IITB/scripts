'''
Script to parse constituency wise data reports (usually report nos. 32, 33).
Pass .xlsx files as input - ECI provides .xls files, must be converted to .xlsx files using Excel to use with this script.

[MODIFIED to output 2024-compliant JSON structure directly and fix data errors]
'''

from openpyxl import load_workbook
import json
from enum import Enum
import string
import os
import re
from collections import defaultdict

# enum to track section of the summary report being parsed currently
class CurrentSection(Enum):
    STATE_UT = "State/UT"
    SUMMARY_CANDIDATE_STATS = "Summary_Candidate_Stats" # Renamed to avoid conflict
    ELECTORS = "Electors"
    VOTERS = "Voters"
    VOTES = "Votes"
    POLLING_STATION = "Polling_Station"
    DATES = "Dates"
    RESULT = "Result"
    NONE = "None"

# --------------------------------------------------------------------------
# 2024-Compliant Template Helpers
# --------------------------------------------------------------------------
def get_empty_gender_obj(default_val=0):
    """Creates a standard gender object, using 0 for electors and None for voters."""
    val = 0 if default_val == 0 else None
    return {"Men": val, "Women": val, "Third_Gender": 0, "Total": default_val}

def get_2024_voters_template():
    """Creates a blank, 2024-compliant Voters object."""
    # For 2024, voter gender data is 'None' (not available), while electors data is '0'.
    # The 'Total' is initialized to 0 and will be parsed.
    return {
        "General": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "OverSeas": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "Proxy": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "Postal": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "Total": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "Votes Not Counted From CU(s) as Per ECI Instructions": {"Men": None, "Women": None, "Third_Gender": 0, "Total": 0},
        "POLLING PERCENTAGE": {"Men": None, "Women": None, "Third_Gender": None, "Total": 0.0}
    }

def get_2024_electors_template():
    """Creates a blank, 2024-compliant Electors object."""
    return {
        "General": get_empty_gender_obj(0),
        "OverSeas": get_empty_gender_obj(0), # Add 'OverSeas'
        "Service": get_empty_gender_obj(0),
        "Total": get_empty_gender_obj(0)
    }

def get_2024_votes_template():
    """Creates a blank, 2024-compliant Votes object."""
    return {
        "Total Votes Polled On EVM": 0,
        "Total Deducted Votes From EVM": 0,
        "Total Valid Votes polled on EVM": 0,
        "Postal Votes Counted": 0,
        "Postal Votes Deducted": 0,
        "Valid Postal Votes": 0,
        "Total Valid Votes Polled": 0,
        "Test Votes polled On EVM": 0,
        "Votes Polled for 'NOTA'(Including Postal)": 0,
        "Tendered Votes": 0,
    }
# --------------------------------------------------------------------------

def clean_value(value):
    """Cleans a single cell value during parsing."""
    if isinstance(value, str):
        value = value.replace(u'\xa0', ' ').strip()
        if value.startswith("=(") and value.endswith(")"):
            try:
                return int(value[2:-1])
            except ValueError:
                pass 
    return value

'''
Function to format constituency names
'''
# --- [CRITICAL FIX 1: Constituency Name Parsing] ---
def format_constituency_name(name):
    """
    Cleans and formats constituency names.
    Removes (SC)/(ST), -SC/-ST, -Gen and trailing numbers.
    """
    if not name or not isinstance(name, str):
        return "Unknown"
    
    name_cleaned = name.strip()
    
    # 1. Remove (SC) or (ST) from anywhere in the string
    name_cleaned = re.sub(r"\s*\((SC|ST)\)\s*", " ", name_cleaned, flags=re.I).strip()
    
    # 2. Remove -SC or -ST suffixes (with optional numbers)
    name_cleaned = re.sub(r"-(SC|ST)-?\d*$", "", name_cleaned, flags=re.I)
    
    # 3. Remove trailing numbers (e.g., -1, -18)
    name_cleaned = re.sub(r"\s*-\s*\d+\s*$", "", name_cleaned)
    
    # 4. Remove standalone -Gen
    name_cleaned = re.sub(r"-Gen$", "", name_cleaned, flags=re.I)
    
    # 5. Capitalize
    parts = name_cleaned.split('-')
    parts = [string.capwords(part.strip()) for part in parts if part.strip()]
    return '-'.join(parts).replace('&', 'and')
# --- [END FIX] ---


# --- FIX: Dictionary to standardize inconsistent state names (keys are lowercase) ---
STATE_NAME_CORRECTIONS = {
    "andhra prade": "Andhra Pradesh",
    "orissa": "Odisha",
    "chhattisgarh": "Chhattisgarh",
    "nct of delhi": "NCT OF Delhi", # Use the canonical casing
    "telangana": "Telangana"
}

# --- [HELPER] ---
def safe_int(value):
    """Safely convert value to integer, handling None, formulas, and commas."""
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '')
        if value.startswith('(') and value.endswith(')'): # Handle "(0)"
            value = value[1:-1]
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0

def safe_float(value):
    """Safely convert value to float, handling None, formulas, and commas."""
    if isinstance(value, (float, int)):
        return float(value)
    if isinstance(value, str):
        value = str(value).strip().replace(',', '').replace('=', '')
        if value.startswith('(') and value.endswith(')'): # Handle "(0)"
            value = value[1:-1]
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0
# --- [END HELPER] ---

'''
Function to parse a sheet in the summary report, each sheet corresponds to one constituency.
'''
def parse_summary_sheet(sheet, year):
    # --- [SCHEMA FIX] ---
    data = {
        "ID": sheet.title.strip(),
        "Constituency": None,
        "State_UT": None,
        "Category": None,
        "Candidates": [], # Must be a list
        "Summary_Candidate_Stats": {}, # Temp key
        "Electors": get_2024_electors_template(),
        "Voters": get_2024_voters_template(),
        "Votes": get_2024_votes_template(),
        "Polling_Station": {"Number": 0, "Average Electors Per Polling": 0},
        "Dates": [],
        "Result": {
            "Winner": {"Party": None, "Candidates": None, "Votes": 0},
            "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0},
            "Margin": 0,
        }
    }
    # --- [END SCHEMA FIX] ---
    
'''
Function to parse a sheet in the summary report, each sheet corresponds to one constituency.
'''
def parse_summary_sheet(sheet, year):
    # --- [SCHEMA FIX] ---
    data = {
        "ID": sheet.title.strip(),
        "Constituency": None,
        "State_UT": None,
        "Category": None,
        "Candidates": [], # Must be a list
        "Summary_Candidate_Stats": {}, # Temp key
        "Electors": get_2024_electors_template(),
        "Voters": get_2024_voters_template(),
        "Votes": get_2024_votes_template(),
        "Polling_Station": {"Number": 0, "Average Electors Per Polling": 0},
        "Dates": [],
        "Result": {
            "Winner": {"Party": None, "Candidates": None, "Votes": 0},
            "Runner-Up": {"Party": None, "Candidates": None, "Votes": 0},
            "Margin": 0,
        }
    }
    # --- [END SCHEMA FIX] ---
    
    current_section = CurrentSection.NONE
    
    for row in sheet.iter_rows(values_only=True):
        if not any(row):
            continue 
    
        cell_one = str(clean_value(row[0]))
        
        # detect/assign sections
        if "State/UT" in cell_one:
            current_section = CurrentSection.STATE_UT
            
            state_name = clean_value(row[1]).split('-')[0].strip()
            # --- [FIX] Case-insensitive lookup ---
            data["State_UT"] = STATE_NAME_CORRECTIONS.get(state_name.lower(), state_name)
            
            # --- [CRITICAL FIX for Summary Parser] ---
            const_raw = str(clean_value(row[3])).strip()
            
            # 1. Extract Category
            cat_match = re.search(r"\((SC|ST)\)", const_raw, re.I)
            if not cat_match:
                # Try the Aruku-ST-1 format
                cat_match = re.search(r"-(SC|ST)", const_raw, re.I)

            if cat_match:
                data["Category"] = cat_match.group(1).upper()
            else:
                data["Category"] = "GENERAL"

            # 2. Clean Constituency Name (use the function)
            data["Constituency"] = format_constituency_name(const_raw)
            # --- [END FIX] ---
            
        elif "CANDIDATES" in cell_one:
            current_section = CurrentSection.SUMMARY_CANDIDATE_STATS
            continue
        elif "ELECTORS" in cell_one:
            current_section = CurrentSection.ELECTORS
            continue
        elif "VOTERS" in cell_one:
            current_section = CurrentSection.VOTERS
            continue
        elif "VOTES" in cell_one:
            current_section = CurrentSection.VOTES
            continue
        elif "POLLING STATION" in cell_one:
            current_section = CurrentSection.POLLING_STATION
            continue
        elif "DATES" in cell_one:
            current_section = CurrentSection.DATES
            # --- [SCHEMA FIX] ---
            poll_date = str(clean_value(row[3]))
            decl_date = str(clean_value(row[5]))
            if poll_date and "polling" not in poll_date.lower():
                data["Dates"].append(poll_date)
            if decl_date and "declaration" not in decl_date.lower():
                data["Dates"].append(decl_date)
            # --- [END SCHEMA FIX] ---
        elif "RESULT" in cell_one:
            current_section = CurrentSection.RESULT
            continue
        
        # parse sections
        if current_section == CurrentSection.SUMMARY_CANDIDATE_STATS or current_section == CurrentSection.ELECTORS:
            key = str(clean_value(row[1]))
            if key and len(row) > 6: # Check key is not None
                third_gender_val = clean_value(row[5])
                data[current_section.value][key] = {
                    "Men": safe_int(row[3]), 
                    "Women": safe_int(row[4]), 
                    "Third_Gender": safe_int(third_gender_val), 
                    "Total": safe_int(row[6])
                }
        elif current_section == CurrentSection.VOTERS:
            key = str(clean_value(row[1]))
            if not key: continue
            
            if "POLLING PERCENTAGE" in key:
                # 2019 files have % in col 3, 2014 has it in col 6
                pct_val = row[3] if row[3] else row[6]
                data["Voters"]["POLLING PERCENTAGE"]["Total"] = safe_float(pct_val)
            
            # --- [FIX for 2024 Voter Gender Data] ---
            elif year == 2024:
                # 2024 report only has Total in row[3] (col D) for Voters.
                # Men, Women, TG data is not provided for voter turnout.
                if key in data[current_section.value]: # Check if key (e.g., "General") exists in the template
                    data[current_section.value][key] = {
                        "Men": None, 
                        "Women": None, 
                        "Third_Gender": 0, # TG is consistently 0 or not provided, so 0 is fine.
                        "Total": safe_int(row[3]) # Total is in Col D
                    }
            # --- [End Fix] ---
            
            elif len(row) > 6: # Existing logic for 2019, 2014
                # This is for General, OverSeas, Postal, etc.
                if key in data[current_section.value]: # Check key exists
                    data[current_section.value][key] = {
                        "Men": safe_int(row[3]), 
                        "Women": safe_int(row[4]), 
                        "Third_Gender": safe_int(row[5]), 
                        "Total": safe_int(row[6])
                    }
                
        elif current_section == CurrentSection.VOTES:
            key = str(clean_value(row[1]))
            if key and len(row) > 6:
                if key in data[current_section.value]:
                    data[current_section.value][key] = safe_int(row[6])
                
        elif current_section == CurrentSection.POLLING_STATION:
            key = str(clean_value(row[1]))
            if key == "Number":
                data[current_section.value][key] = safe_int(row[3])
            elif "Average Electors" in key:
                 data[current_section.value]["Average Electors Per Polling"] = safe_int(row[6])
                 
        elif current_section == CurrentSection.RESULT:
            # This data is unreliable, but we parse it anyway
            key = str(clean_value(row[1]))
            if key and len(row) > 6:
                if key != "Margin":
                    data[current_section.value][key] = {
                        "Party": clean_value(row[3]), 
                        "Candidates": clean_value(row[4]), 
                        "Votes": safe_int(row[6])
                    }
                else:
                    data[current_section.value][key] = safe_int(row[3])
                    current_section = CurrentSection.NONE

    return data

'''
Function to parse the detailed result report which contains candidate-wise data for each constituency.
Each row corresponds to one candidate.
'''
# --- [CRITICAL FIX for Detailed Parser] ---
def parse_detailed_sheet(sheet, ids, year, header_map):
    candidates = defaultdict(list)
    rows = list(sheet.iter_rows(values_only=True))
    
    if year <= 2014: 
        header_row_index = 0
        subheader_row_index = 1
        data_start_row = 2
    else: # 2019 and 2024
        header_row_index = 1
        subheader_row_index = 2
        data_start_row = 3
        
    l1_fields = [str(clean_value(h)).replace('\n', ' ').strip().lower() if h else "" for h in rows[header_row_index]]
    l2_fields = [str(clean_value(h)).replace('\n', ' ').strip().lower() if h else "" for h in rows[subheader_row_index]]

    # Fix merged cells in L1 headers
    last_valid_header = ""
    for i in range(len(l1_fields)):
        if l1_fields[i]:
            last_valid_header = l1_fields[i]
        else:
            l1_fields[i] = last_valid_header

    # --- Create Efficient, Case-Insensitive Lookup Map ---
    constituency_lookup = {
        (v['State_UT'].lower(), v['Constituency'].lower()): k
        for k, v in ids.items()
        if v['State_UT'] and v['Constituency']
    }
    
    failed_states = set()

    # --- [NEW HEADER LOGIC] ---
    for i in range(len(l2_fields)):
        l1 = l1_fields[i]
        l2 = l2_fields[i]

        # Check L2 first (most specific)
        if "candidate" in l2: # Match 'candidate name' or 'candidates name'
            header_map["Candidate Name"] = i
        elif l2 == "sex": # Exact match for "SEX"
            header_map["Gender"] = i
        elif l2 == "age": # Exact match for "AGE"
            header_map["Age"] = i
        elif l2 == "category": # Exact match for "CATEGORY"
            header_map["Category"] = i
        elif "party name" in l2:
            header_map["Party Name"] = i
        elif "party symbol" in l2:
            header_map["Party Symbol"] = i
        
        # Check combined headers for ambiguous fields
        elif "votes secured" in l1 and l2 == "general":
            header_map["General"] = i
        elif "votes secured" in l1 and l2 == "postal":
            header_map["Postal"] = i
        elif "votes secured" in l1 and l2 == "total":
            header_map["Total"] = i
        
        # This is for 2019/2024.
        elif "% of votes secured" in l1 and "over total electors" in l2:
             header_map["% Over Total Electors"] = i
        elif "% of votes secured" in l1 and "over total votes polled" in l2:
             header_map["% Over Total Votes Polled"] = i
        elif "% of votes secured" in l1 and "over total valid votes" in l2:
             header_map["% Over Total Valid Votes"] = i
        
        # Check L1 or L2 for Total Electors
        # Use 'in' for flexibility (e.g. 'total electors' vs 'total electors in constituency')
        elif "total electors" in l1 or "total electors" in l2:
             header_map["Total Electors"] = i

    # --- [FALLBACK for 2014] ---
    if year <= 2014:
        # 2014 has a different header structure
        for i in range(len(l2_fields)):
             l1 = l1_fields[i]
             l2 = l2_fields[i]
             if "candidate" in l2 and "name" in l2: header_map["Candidate Name"] = i
             elif l2 == "sex": header_map["Gender"] = i
             elif l2 == "age": header_map["Age"] = i
             elif l2 == "category": header_map["Category"] = i
             elif "party name" in l2: header_map["Party Name"] = i
             elif "party symbol" in l2: header_map["Party Symbol"] = i
             elif l1 == "votes secured" and l2 == "general": header_map["General"] = i
             elif l1 == "votes secured" and l2 == "postal": header_map["Postal"] = i
             elif l1 == "votes secured" and l2 == "total": header_map["Total"] = i
             elif "over total electors" in l2: header_map["% Over Total Electors"] = i
             elif "over total votes polled" in l2: header_map["% Over Total Votes Polled"] = i
             elif l2 == "total electors": header_map["Total Electors"] = i
        
        # 2014 does not have "% Over Total Valid Votes", it's calculated in the merge step
        header_map["% Over Total Valid Votes"] = -2 # Mark as "to be calculated"
    # --- [END 2014 FALLBACK] ---
    
    # For 2019, if '% Over Total Valid Votes' is not present, mark it for calculation
    if year > 2014 and header_map["% Over Total Valid Votes"] == -1:
        print("  -> Note: '% Over Total Valid Votes' not found, will be calculated.")
        header_map["% Over Total Valid Votes"] = -2 # Mark as "to be calculated"
    # --- [END NEW HEADER LOGIC] ---

    missing_keys = [k for k,v in header_map.items() if v == -1] # Check for any unmapped keys
    if missing_keys:
        print(f"  -> WARNING: Could not find headers in detailed sheet: {missing_keys}")
        print(f"     L1 HEADERS: {l1_fields}")
        print(f"     L2 HEADERS: {l2_fields}")
        
    # --- Process data rows ---
    for row in rows[data_start_row:]:
        # Use a safe-check for candidate name column
        if header_map["Candidate Name"] == -1 or not row[header_map["Candidate Name"]]:
            continue

        state = clean_value(row[0])
        state_standardized = STATE_NAME_CORRECTIONS.get(state.lower(), state)
            
        constituency = format_constituency_name(clean_value(row[1]))
        
        lookup_key = (state_standardized.lower(), constituency.lower())
        constituency_id = constituency_lookup.get(lookup_key)
        
        if not constituency_id:
            if state.lower() != "state name" and state not in failed_states:
                print(f'  -> Lookup failed for: (State: "{state}", Standardized: "{state_standardized}", Const: "{constituency}")')
                failed_states.add(state)
            continue # Skip rows we can't map

        try:
            # Get the value for '% Over Total Valid Votes'. If column doesn't exist, default to 0.0
            pct_valid_votes = 0.0
            if header_map["% Over Total Valid Votes"] >= 0: # Check if index was found
                pct_valid_votes = round(safe_float(row[header_map["% Over Total Valid Votes"]]), 2)
            
            # Build the 2024-compliant object *manually*
            candidate_data = {
                "Candidate Name": clean_value(row[header_map["Candidate Name"]]),
                "Gender": clean_value(row[header_map["Gender"]]),
                "Age": safe_int(row[header_map["Age"]]),
                "Category": clean_value(row[header_map["Category"]]),
                "Party Name": clean_value(row[header_map["Party Name"]]),
                "Party Symbol": clean_value(row[header_map["Party Symbol"]]),
                "Total Votes Polled In The Constituency": 0, # Placeholder
                "Valid Votes": 0, # Placeholder
                "Votes Secured": {
                    "General": safe_int(row[header_map["General"]]),
                    "Postal": safe_int(row[header_map["Postal"]]),
                    "Total": safe_int(row[header_map["Total"]])
                },
                "% of Votes Secured": {
                    "Over Total Electors In Constituency": round(safe_float(row[header_map["% Over Total Electors"]]), 2),
                    "Over Total Votes Polled In Constituency": round(safe_float(row[header_map["% Over Total Votes Polled"]]), 2)
                },
                
                "Over Total Valid Votes Polled In Constituency": pct_valid_votes, # Use the value or 0.0
                "Total Electors": safe_int(row[header_map["Total Electors"]])
            }
            candidates[constituency_id].append(candidate_data)
        except Exception as e:
            print(f"Error processing candidate row for {constituency_id}: {e}")
            print(f"Row data: {row}")

    if failed_states:
        print(f"  -> WARNING: Lookup failed for {len(failed_states)} unique state names.")
        print("     Please add these names to the STATE_NAME_CORRECTIONS dictionary if they are not just headers.")
        
    return candidates
# --- [END CRITICAL FIX] ---


JOBS_CONFIG = [
    {
        "year": 2024,
        "summary_path": 'converted_xlsx_reports/election_data_2024_xlsx/32-Constituency-Data-Summery-Report.xlsx',
        "detailed_path": 'converted_xlsx_reports/election_data_2024_xlsx/33-Constituency-Wise-Detailed-Result.xlsx',
        "output_path": 'parsed/lundori2024.json'
    },
    {
        "year": 2019,
        "summary_path": 'converted_xlsx_reports/election_data_2019_xlsx/32. Constituency Data Summary Report.xlsx',
        "detailed_path": 'converted_xlsx_reports/election_data_2019_xlsx/33. Constituency Wise Detailed Result.xlsx',
        "output_path": 'parsed/lundori2019.json'
    },
    # {
    #     "year": 2014,
    #     "summary_path": 'converted_xlsx_reports/election_data_2014_xlsx/Constituency data summary.xlsx',
    #     "detailed_path": 'converted_xlsx_reports/election_data_2014_xlsx/Constituency wise detailed result.xlsx',
    #     "output_path": 'parsed/parsed2014.json'
    # },
]

for job in JOBS_CONFIG:
    year = job["year"]
    summary_file_path = job["summary_path"]
    detailed_file_path = job["detailed_path"]
    out_path = job["output_path"]

    print(f"\n--- Starting processing for year: {year} ---")

    try:
        wb = load_workbook(summary_file_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: File not found. Skipping year {year}.")
        print(f"Path: {summary_file_path}")
        continue

    parsed = []
    for s in wb.sheetnames:
        sheet = wb[s]
        parsed.append(parse_summary_sheet(sheet, year))
    print(f"Parsed {len(parsed)} constituency summaries.")

    ids = {}
    for c in parsed:
        if c['State_UT'] and c['Constituency']:
            ids[c['ID']] = {'State_UT': c['State_UT'], 'Constituency': c['Constituency']}
        else:
            print(f"Warning: Skipping sheet '{c['ID']}' due to missing State or Constituency.")
    print(f"Created lookup map with {len(ids)} entries.")

    if year == 2014: # Left this debug block, it's helpful
        print("\n--- DEBUG: 2014 State/Constituency pairs from Summary ---")
        sample_keys = set()
        for v in ids.values():
            if v['State_UT'] and v['Constituency']:
                sample_keys.add((v.get('State_UT'), v.get('Constituency'))) 
        
        printed_states = set()
        for state, const in sample_keys:
            if state and state not in printed_states:
                print(f"  -> Found in Summary: (State: \"{state}\", Constituency: \"{const}\")")
                printed_states.add(state)
        print("----------------------------------------------------------\n")

    try:
        wb = load_workbook(detailed_file_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: File not found. Skipping year {year}.")
        print(f"Path: {detailed_file_path}")
        continue
    
    # --- [NAMEERROR FIX] ---
    # Initialize header_map here, in the loop's scope
    header_map = {
        "Candidate Name": -1, "Gender": -1, "Age": -1, "Category": -1, 
        "Party Name": -1, "Party Symbol": -1, "General": -1, "Postal": -1, 
        "Total": -1, "% Over Total Electors": -1, "% Over Total Votes Polled": -1,
        "Total Electors": -1, "% Over Total Valid Votes": -1
    }
    # Pass it to the function
    candidates = parse_detailed_sheet(wb.active, ids, year, header_map)
    # --- [END NAMEERROR FIX] ---
    
    print(f"Parsed candidate data for {len(candidates)} constituencies.")

    print("Merging candidate data into summary...")
    for constituency in parsed:
        full_id = constituency['ID']
        if full_id in candidates:
            candidate_list = candidates[full_id]
            
            # Get summary vote data
            total_polled = constituency.get("Voters", {}).get("Total", {}).get("Total", 0)
            valid_votes = constituency.get("Votes", {}).get("Total Valid Votes Polled", 0)
            
            # Add summary data to each candidate
            for cand in candidate_list:
                cand["Total Votes Polled In The Constituency"] = total_polled
                cand["Valid Votes"] = valid_votes
                
                # --- [DATA FIX] ---
                # If % over valid votes was not in the sheet (e.g. 2014, or 2019 fallback), calculate it.
                if (header_map["% Over Total Valid Votes"] == -2) and valid_votes > 0:
                    cand["Over Total Valid Votes Polled In Constituency"] = round(
                        (cand["Votes Secured"]["Total"] / valid_votes) * 100, 2
                    )
                # --- [END DATA FIX] ---
            
            constituency['Candidates'] = candidate_list
            
            # --- [DATA FIX] Recalculate Result ---
            if candidate_list: # Only if we have candidates
                try:
                    # Sort candidates by total votes, descending
                    sorted_candidates = sorted(
                        candidate_list, 
                        key=lambda c: c["Votes Secured"]["Total"], 
                        reverse=True
                    )
                    
                    # Update Winner
                    if len(sorted_candidates) > 0:
                        winner = sorted_candidates[0]
                        constituency["Result"]["Winner"] = {
                            "Party": winner["Party Name"],
                            "Candidates": winner["Candidate Name"],
                            "Votes": winner["Votes Secured"]["Total"]
                        }
                    
                    # Update Runner-Up
                    if len(sorted_candidates) > 1:
                        runner_up = sorted_candidates[1]
                        constituency["Result"]["Runner-Up"] = {
                            "Party": runner_up["Party Name"],
                            "Candidates": runner_up["Candidate Name"],
                            "Votes": runner_up["Votes Secured"]["Total"]
                        }
                        # Update Margin
                        constituency["Result"]["Margin"] = (
                            winner["Votes Secured"]["Total"] - runner_up["Votes Secured"]["Total"]
                        )
                    elif len(sorted_candidates) > 0: # Only a winner
                         constituency["Result"]["Runner-Up"] = {"Party": None, "Candidates": None, "Votes": 0}
                         constituency["Result"]["Margin"] = winner["Votes Secured"]["Total"]

                except Exception as e:
                    print(f"Error calculating winner for {full_id}: {e}")
            # --- [END DATA FIX] ---
            
        else:
            print(f"Warning: No candidate data found for {constituency['ID']} ({constituency['Constituency']})")
            constituency['Candidates'] = [] # Ensure it's an empty list

        # --- [SCHEMA FIX] Remove Summary_Candidate_Stats ---
        if "Summary_Candidate_Stats" in constituency:
            del constituency["Summary_Candidate_Stats"]

    output_dir = os.path.dirname(out_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    print(f"Writing final JSON to: {out_path}")
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(parsed, f, indent=4, default=str)

    print(f"--- Finished processing for year: {year} ---")

print("\n✅ All jobs complete.")