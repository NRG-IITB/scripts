import openpyxl
import os

def delete_columns_from_workbook(input_filepath, output_filepath):
    """
    Opens an .xlsx workbook, deletes columns E, G, and I from every sheet,
    and saves the modified workbook to a new file.

    Args:
        input_filepath (str): The path to the original Excel file.
        output_filepath (str): The path to save the new Excel file.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(input_filepath)
        print(f"Successfully loaded '{input_filepath}'.")

        # Get all sheet names
        sheet_names = wb.sheetnames
        print(f"Found sheets: {', '.join(sheet_names)}")

        # Columns to delete (E=5, G=7, I=9)
        # We must delete them in reverse order (I, G, E) so that
        # the column indices don't shift unexpectedly.
        cols_to_delete = [9, 7, 5]
        cols_to_delete_names = ['I', 'G', 'E']

        # Check if there is more than one sheet before skipping
        if len(sheet_names) > 1:
            sheets_to_process = sheet_names[1:] # Get all sheets except the first one
            print(f"Skipping first sheet: '{sheet_names[0]}'")
        else:
            sheets_to_process = [] # No sheets to process if there's only one
            print("Only one sheet found, nothing to process after skipping the first sheet.")

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            print(f"  Processing sheet: '{sheet_name}'...")
            
            for col_idx, col_name in zip(cols_to_delete, cols_to_delete_names):
                try:
                    ws.delete_cols(col_idx, 1) # (column_index, number_of_columns_to_delete)
                    print(f"    - Deleted column {col_name} (index {col_idx}).")
                except Exception as e:
                    print(f"    - Could not delete column {col_name} (index {col_idx}): {e}")

        # Save the modified workbook to the new file path
        wb.save(output_filepath)
        print(f"\nSuccessfully processed all sheets and saved to '{output_filepath}'.")

    except FileNotFoundError:
        print(f"Error: The file '{input_filepath}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # --- Configuration ---
    # Set the path to your original file
    original_file = "2014_summary.xlsx"
    # Set the path for the new, modified file
    new_file = "2014_summary_new.xlsx"
    # ---------------------

    # --- Create a dummy file for testing if it doesn't exist ---
    if not os.path.exists(original_file):
        print(f"'{original_file}' not found. Creating a dummy file for testing.")
        try:
            dummy_wb = openpyxl.Workbook()
            ws1 = dummy_wb.active
            ws1.title = "Sheet1"
            # Populate some data
            for r in range(1, 10):
                for c in range(1, 15):
                    col_letter = openpyxl.utils.get_column_letter(c)
                    ws1.cell(row=r, column=c, value=f"{col_letter}{r}")
            
            # Add a second sheet
            ws2 = dummy_wb.create_sheet("Sheet2")
            for r in range(1, 10):
                for c in range(1, 15):
                    col_letter = openpyxl.utils.get_column_letter(c)
                    ws2.cell(row=r, column=c, value=f"{col_letter}{r}")

            dummy_wb.save(original_file)
            print(f"Dummy file '{original_file}' created with two sheets.")
        except Exception as e:
            print(f"Could not create dummy file: {e}")
    # -----------------------------------------------------------

    # Run the function
    delete_columns_from_workbook(original_file, new_file)